"""
@author:   Ken Venner
@contact:  ken@venerllc.com
@version: 1.43

Library of tools used to process XLS/XLSX files
"""

import openpyxl  # xlsx (read/write)

# comment out below if we are XLSX ONLY
import xlrd  # xls (read)
import xlwt  # xls (write)
from xlutils.copy import (
    copy as xl_copy,
)  # xls(read copy over tool to enalve write)/ pip install xlutils

import os  # determine if a file exists
import pprint
import json

from dataclasses import dataclass, fields
from typing import List, Any, Tuple
from types import NoneType

import kvdate
import kvmatch
import datetime
import re

# logging
import logging

logger = logging.getLogger(__name__)

# global variables
AppVersion = "1.43"

# set to true in kvxlsx.py
XLSXONLY = False

# ----- OPTIONS ---------------------------------------
# debug
# dupkeyfail
# data_only
# req_cols
# col_aref
# xlat_dict
# dictkeys
#
# optiondict:
# col_header
# no_header
# aref_result
# save_row
# save_row_abs
# save_col_abs
# save_colmap
# save_colfmt = "column_name"
# start_row
# sheetname
# sheetrow
# dateflds

# -- CONSTANTS -- #
ILLEGAL_CHARACTERS_RE_STR_ORIG = r"[\000-\010]|[\013-\014]|[\016-\037]"

ILLEGAL_CHARACTERS_LIST = (
    [chr(x) for x in range(0, 8 + 1)]
    + [chr(x) for x in range(11, 12 + 1)]
    + [chr(x) for x in range(14, 15 + 1)]
)
ILLEGAL_CHARACTERS_STR = "".join(ILLEGAL_CHARACTERS_LIST)

ILLEGAL_CHARACTERS_RE = r"|".join(ILLEGAL_CHARACTERS_LIST)

ILLEGAL_CHARACTERS_TRANS_TBL = str.maketrans(
    ILLEGAL_CHARACTERS_STR, " " * len(ILLEGAL_CHARACTERS_STR)
)


if False:
    print(f"{ILLEGAL_CHARACTERS_LIST=}")
    print(f"{ILLEGAL_CHARACTERS_RE=}")
    print(f"{ILLEGAL_CHARACTERS_STR=}")

VALID_XLSX_STYLE_UNDERLINE = (
    "single",  # single underline
    "double",  # double underline
    "singleAccounting",  # single accounting underline
    "doubleAccounting",  # double accounting underline
    "None",  # no underline
)


FLD_XLSROW = "XLSRow"
FLD_XLSROW_ABS = "XLSRowAbs"
FLD_XLSCOL_ABS = "XLSColAbs1"
FLD_XLSNEW_COLMAP = "XLSColMap"
FLD_XLSFMT = "XLSFmt"


@dataclass
class ExcelConfig:
    xlsxfiletype: bool | None = (
        None  # true when the file tupe is XLSX otherwise false
    )
    filename: str | None = None  # the filename this configuration is related to

    req_cols: list | None = (
        None  # list of column headers we must find to say we found the header
    )
    col_aref: list | None = (
        None  # list of column headers we define/override it to be
    )
    xlatdict: dict | None = (
        None  # file value to desired value column heading mapping
    )

    allow_empty: bool = False  # if true - we allow a header to be read in with no data - otherwise we raise and exceptoin
    aref_result: bool = (
        False  # if true - we don't return dicts, we return a list
    )
    col_header: bool = False  # if true - we take the first row of the file as the header, we don't go looking for the header
    data_only: bool = True  # if true - we open the file as data_only
    keep_vba: bool = True  # if true - then load the xlsx with vba scripts on and save as xlsm
    no_header: bool = (
        False  # if true - there are no headers read - we either return
    )
    no_warnings: bool = False  # if true - do not display warning message

    sheetname: str | None = None  # sheet name we are processing
    sheetrow: int | None = None  # index number of sheet you want

    start_row: int = 0  # if passed in - we start the search at this row (starts at 1 or greater)
    # change user input to reduce by one as we are zero based - so if we start on row 1 - this value should be zero
    max_rows: int = 100000000  # max rows we process to before giving up

    row_header: int | None = None

    save_row: bool = (
        False  # if true - then we append/save the XLSRow with the record
    )
    save_row_abs: bool = False  # if true - then we append/save the openpxl row (only works on XLSX)
    save_col_abs: bool = False  # if true - then we append/save the absolute xlsx column number of the first column - from openpyxl (only works on XLSX)
    save_col_fmt: bool = False  # if populated with a column header - than we capture and save the format of that column

    replace_sheet: bool = False  # when true, we are adding/overridding a sheet into an exising file if one exists or creating the file
    replace_index: int | None = (
        None  # if we want to position the new sheet- we can define where we want it
    )


FLDS_EXCELCONFIG = [f.name for f in fields(ExcelConfig)]

# create the list of misconfigured solutions
BADOPTIONDICT = {
    "allowempty": "allow_empty",
    "header_empty": "allow_empty",
    "headerempty": "allow_empty",
    "aref_results": "aref_result",
    "arefresult": "aref_result",
    "arefresults": "aref_result",
    "col_headers": "col_header",
    "colheader": "col_header",
    "colheaders": "col_header",
    "keepvba": "keep_vba",
    "max_row": "max_rows",
    "maxrow": "max_rows",
    "maxrows": "max_rows",
    "no_headers": "no_header",
    "noheader": "no_header",
    "noheaders": "no_header",
    "replaceindex": "replace_index",
    "addsheet": "replace_sheet",
    "addsheets": "replace_sheet",
    "add_sheet": "replace_sheet",
    "add_sheets": "replace_sheet",
    "replacesheet": "replace_sheet",
    "replacesheets": "replace_sheet",
    "replace_sheets": "replace_sheet",
    "save_cols_abs": "save_col_abs",
    "save_colsabs": "save_col_abs",
    "savecolabs": "save_col_abs",
    "savecolsabs": "save_col_abs",
    "save_col_fmt": "save_colfmt",
    "savecol_fmt": "save_colfmt",
    "savecolfmt": "save_colfmt",
    "savecolmap": "save_colmap",
    "save_rows_abs": "save_row_abs",
    "save_rowsabs": "save_row_abs",
    "saverowabs": "save_row_abs",
    "saverowsabs": "save_row_abs",
    "save_rows": "save_row",
    "saverow": "save_row",
    "saverows": "save_row",
    "sheet_name": "sheetname",
    "sheetrows": "sheetrow",
    "sheet_row": "sheetrow",
    "sheet_rows": "sheetrow",
    "startrow": "start_row",
    "startrows": "start_row",
    "start_rows": "start_row",
}


# ---- UTILITY FUNCTIONS ------------------------------


def create_excel_config(
    optiondict: dict | None = None,
    func_name: str | None = None,
    debug: bool = False,
) -> ExcelConfig:
    """
    Create an exceldict configuratoin object using default values from optoindict

    Inputs:
        optiondict - dict - configuratoin options
        func_name - str - if you want to display the name of the funciton when remapping optiondict values
        debug - bool - when true, we display status messages

    Returns:
        ExcelConfig data class

    """

    # set inputs if not set
    if optiondict is None:
        optiondict = dict()
    # test inputs
    if not isinstance(optiondict, dict):
        raise TypeError(f"optiondict must be dict but is: {type(optiondict)}")
    if func_name is None:
        func_name = ""
    if not isinstance(func_name, str):
        raise TypeError(f"func_name must be str but is: {type(func_name)}")

    # run through bad optoins to fix bad keys
    msg = kvmatch.badoptiondict_check(
        func_name,
        optiondict,
        BADOPTIONDICT,
        noshowwarning=True,
        fix_missing=True,
    )

    if debug:
        print(f"{msg=}")

    # pull from option dict all the values that we are going to set for configuration
    cfg_dict = {k: v for k, v in optiondict.items() if k in FLDS_EXCELCONFIG}

    if debug:
        print(f"{cfg_dict=}")

    # if they passed in start_row we need to decrement by one
    # as humans start with 1 and this tool starts with zero
    if "start_row" in cfg_dict:
        cfg_dict["start_row"] -= 1

    # create the configuration object
    excel_config = ExcelConfig(**cfg_dict)

    return excel_config


def strip_xls_illegal_chars(value: str, debug: bool = False) -> str:
    """
    Remove characters from string that can not be placed in XLSX files

    Input:
        value - str/byte - the string to be changed
        debug - bool - when true, display messages

    Return:
        value_new - str/byte - the string with illegal characters repalced with a space

    """
    if isinstance(value, (str, bytes)):
        if False:
            newvalue = re.sub(ILLEGAL_CHARACTERS_RE, " ", value)
            if debug:
                print(f"{value=}")
                print(f"{newvalue=}")
            return newvalue
        else:
            newvalue = value.translate(ILLEGAL_CHARACTERS_TRANS_TBL)
            if debug:
                print(f"{value=}")
                print(f"{newvalue=}")
            return newvalue

    else:
        return value


def xldate_to_datetime(
    xldate: str | int | float, skipblank: bool = False
) -> datetime.datetime | Any:
    """
    Convert an XLS date number to a python datetime object

    Input:
        xlsdate - str/int/float - the XLS date as a string or an int that needs to be converted
        skipblank - bool

    Returns:
        xls_datetime - datetime representative of that passed in value

    """

    if isinstance(xldate, str):
        # string - convert using string to date routines
        logger.debug(
            "converting xldate string to date using kvdate.datetime_from_str:%s",
            xldate,
        )
        return kvdate.datetime_from_str(xldate, skipblank)
    elif isinstance(xldate, (int, float)):
        # int - use the defined math to convert
        logger.debug("converting xldate float to date:%s", xldate)
        temp = datetime.datetime(1899, 12, 30)
        delta = datetime.timedelta(days=xldate)
        return temp + delta
    else:
        # when it is neither a string or an int
        logger.warning(
            f"could not convert value it is not [str,int] but is: {type(xldate)}"
        )
        return xldate


def _extract_excel_row_into_list(
    xlsxfiletype: bool,
    s,
    row: int,
    colstart: int,
    colmax: int,
    debug: bool = False,
) -> tuple[List[Any | None], int | None, int | None]:
    """
    Extract a row from an Excel object defined by s
    and pass back a list of these values

    Input:
        xlsxfiletype - bool - when true - this is for an XLSX file, when false thjis is for XLS file
        s
        row
        colstart
        colmax
        debug

    Returns:
        row - list of values extracted from cells

    """

    # debugging
    if debug:
        print("_extract_excel_row_into_list:row:", row)
        print("_extract_excel_row_into_list:xlsxfiletype:", xlsxfiletype)
    logger.debug("row: %s", row)
    logger.debug("xlsxfiletype:%s", xlsxfiletype)

    # capture row and first column
    if xlsxfiletype:
        c_row = s.cell(row=row + 1, column=colstart + 1).row
        c_col = s.cell(row=row + 1, column=colstart + 1).column
    else:
        # c_row = s.cell(row, colstart).row
        # c_col = s.cell(row, colstart).column
        c_row = None
        c_col = None

    # clear the row
    rowdata = []

    # pull each column out of XLS and build the row array
    for col in range(colstart, colmax):
        # get cell value
        if xlsxfiletype:
            c_value = s.cell(row=row + 1, column=col + 1).value
        else:
            c_value = s.cell(row, col).value

        # debugging
        if debug:
            print("row:", row, ":col:", col, ":cValue:", c_value)
        logger.debug("row:%s:col:%s:cValue:%s", row, col, c_value)

        # add this value to the array that will be used to determine if this is header
        rowdata.append(c_value)

    # return the row
    return rowdata, c_row, c_col


def getExcelCellValue(
    excel_dict: dict, row: int, col_name: str, debug: bool = False
) -> Any | None:
    """
    For a defined row number and a column name string, extract the value for that cell

    Inputs:
        excel_dict - custom excel dictionary created to represent a worksheet object
        row - int - row number that we are extracting from
        col_name - str - the header name associated with the column
        debug - bool - when true, we display debugging messages

    Returns:
        cell_value - the value in the cell of interest.


    """
    if debug:
        print("getExcelCellValue:excel_dict:", excel_dict)
        print("getExcelCellValue:row:", row)
        print("getExcelCellValue:col_name:", col_name)
    logger.debug("excel_dict:%s", excel_dict)
    logger.debug("row:%s", row)
    logger.debug("col_name:%s", col_name)

    # determine the col # we are using by doing a header lookup for col_name
    # add to that the offset if the starting column is not the first columnn
    # if the col_name provided was not in the header - it will automatically evoke a ValueError reporting the problem
    # so we will just allow that to be the error message
    col = excel_dict["header"].index(col_name) + excel_dict["sheetmincol"]

    # get cell value
    if excel_dict["xlsxfiletype"]:
        return (
            excel_dict["s"]
            .cell(row=row + 1 + excel_dict["row_header"], column=col + 1)
            .value
        )
    else:
        return excel_dict["s"].cell(row, col).value


def setExcelCellValue(
    excel_dict: dict, row: int, col_name: str, value: Any, debug: bool = False
) -> None:
    """
    For a defined row number and a column name string, set the value for that cell

    Inputs:
        excel_dict - custom excel dictionary created to represent a worksheet object
        row - int - row number that we are extracting from
        col_name - str - the header name associated with the column
        debug - bool - when true, we display debugging messages

    Returns:
        cell_value - the value in the cell of interest.

    """

    if debug:
        print("setExcelCellValue:excel_dict:", excel_dict)
        print("setExcelCellValue:row:", row)
        print("setExcelCellValue:col_name:", col_name)
    logger.debug("excel_dict:%s", excel_dict)
    logger.debug("row:%s", row)
    logger.debug("col_name:%s", col_name)

    # determine the col # we are using by doing a header lookup for col_name
    # add to that the offset if the starting column is not the first columnn
    # if the col_name provided was not in the header - it will automatically evoke a ValueError reporting the problem
    # so we will just allow that to be the error message
    col = excel_dict["header"].index(col_name) + excel_dict["sheetmincol"]

    # get cell value
    if excel_dict["xlsxfiletype"]:
        excel_dict["s"].cell(
            row=row + 1 + excel_dict["row_header"], column=col + 1, value=value
        )
    else:
        logger.error("feature not supported on xls file - only XLSX")
        raise NotImplementedError(
            "feature not supported on xls file - only XLSX"
        )


# routine to get a cell fill pattern - returns the (rgb, solid) values
def getExcelCellFont(
    excel_dict: dict, row: int, col_name: str, debug: bool = False
) -> tuple[Any, Any, Any, Any, Any, Any, Any]:
    """
    For a defined row number and a column name string, get the cell_font information for that cell

    Inputs:
        excel_dict - custom excel dictionary created to represent a worksheet object
        row - int - row number that we are extracting from
        col_name - str - the header name associated with the column
        debug - bool - when true, we display debugging messages

    Returns:
        cell_font_name
        cell_font_size
        cell_font_bold
        cell_font_italic
        cell_font_underline
        cell_font_strike
        cell_font_color

    """
    if debug:
        print("getExcelCellFont:excel_dict:", excel_dict)
        print("getExcelCellFont:row:", row)
        print("getExcelCellFont:col_name:", col_name)
    logger.debug("excel_dict:%s", excel_dict)
    logger.debug("row:%s", row)
    logger.debug("col_name:%s", col_name)

    # determine the col # we are using by doing a header lookup for col_name
    # add to that the offset if the starting column is not the first columnn
    # if the col_name provided was not in the header - it will automatically evoke a ValueError reporting the problem
    # so we will just allow that to be the error message
    col = excel_dict["header"].index(col_name) + excel_dict["sheetmincol"]

    # debugging
    if debug:
        print("cell_font")
        print("col_name:", col_name)
        print("col:", col)
        print("row:", row)
        print(
            "value:",
            excel_dict["s"]
            .cell(row=row + 1 + excel_dict["row_header"], column=col + 1)
            .value,
        )

    # get cell font
    if excel_dict["xlsxfiletype"]:
        # get font settings
        cell_font = (
            excel_dict["s"]
            .cell(row=row + 1 + excel_dict["row_header"], column=col + 1)
            .font
        )
        # debugging
        if debug:
            print(
                "getExcelCellFont:name:",
                cell_font.name,
                "getExcelCellFont:size:",
                cell_font.size,
                "getExcelCellFont:strike:",
                cell_font.strike,
            )

        # return cell_font
        return (
            cell_font.name,
            cell_font.size,
            cell_font.bold,
            cell_font.italic,
            cell_font.underline,
            cell_font.strike,
            cell_font.color,
        )
    else:
        logger.error("feature not supported on xls file - only XLSX")
        raise NotImplementedError(
            "feature not supported on xls file - only XLSX"
        )


# routine to set a cell fill pattern
def setExcelCellFont(
    excel_dict: dict,
    row: int,
    col_name: str,
    name: str | None = None,
    size: str | None = None,
    bold: bool | None = None,
    italic: str | None = None,
    underline: str | None = None,
    strike: str | None = None,
    color: str | None = None,
    debug: bool = False,
) -> None:
    """
    For a defined row number and a column name string, get the cell_font information for that cell

    Inputs:
        excel_dict - custom excel dictionary created to represent a worksheet object
        row - int - row number that we are extracting from
        col_name - str - the header name associated with the column
        debug - bool - when true, we display debugging messages

    Optional Inputs:
        name - str
        size - float
        bold - bool
        italic - bool
        underline - str
            "single"   single underline
            "double"   double underline
            "singleAccounting"   single accounting underline
            "doubleAccounting"   double accounting underline
            None   no underline
        strike
        color - a hex color code, when reading it yuou need to read color.rgb to get the hex value

    """

    if debug:
        print("setExcelCellFont:excel_dict:", excel_dict)
        print("setExcelCellFont:row:", row)
        print("setExcelCellFont:col_name:", col_name)

    # determine the col # we are using by doing a header lookup for col_name
    # add to that the offset if the starting column is not the first columnn
    # if the col_name provided was not in the header - it will automatically evoke a ValueError reporting the problem
    # so we will just allow that to be the error message
    col = excel_dict["header"].index(col_name) + excel_dict["sheetmincol"]

    # get cell value
    if excel_dict["xlsxfiletype"]:
        # get the cell
        cell = excel_dict["s"].cell(
            row=row + 1 + excel_dict["row_header"], column=col + 1
        )

        # get the current settings
        orig_cell_font = cell.font

        # create the new font object
        cell_font_args = {}

        # set the values if they are not None
        if name is not None:
            cell_font_args["name"] = name

        if size is not None:
            cell_font_args["size"] = size

        if bold is not None:
            cell_font_args["bold"] = bold

        if italic is not None:
            cell_font_args["italic"] = italic

        if underline is not None:
            # test it is a valid value
            if underline not in VALID_XLSX_STYLE_UNDERLINE:
                raise ValueError(
                    f"underline value [{underline}] not in valid list: {'|'.join(VALID_XLSX_STYLE_UNDERLINE)}"
                )
            if underline == "None":
                underline = None
            cell_font_args["underline"] = underline

        if strike is not None:
            cell_font_args["strike"] = strike

        if color is not None:
            cell_font_args["color"] = color

        # if anything was populated - then set the cell.font to what was passed in
        if cell_font_args:
            if debug:
                print(f"{cell_font_args=}")
                print(f"{cell.font.__dict__=}")

            new_cell_font = openpyxl.styles.Font(
                name=cell_font_args.get("name", orig_cell_font.name),
                size=cell_font_args.get("size", orig_cell_font.size),
                bold=cell_font_args.get("bold", orig_cell_font.bold),
                italic=cell_font_args.get("italic", orig_cell_font.italic),
                underline=cell_font_args.get(
                    "underline", orig_cell_font.underline
                ),
                strike=cell_font_args.get("strike", orig_cell_font.strike),
                color=cell_font_args.get("color", orig_cell_font.color),
            )

            if debug:
                print(f"{new_cell_font.name=}")
                print(f"{new_cell_font.size=}")
                print(f"{new_cell_font.bold=}")
                print(f"{new_cell_font.italic=}")
                print(f"{new_cell_font.underline=}")
                print(f"{new_cell_font.strike=}")
                print(f"{new_cell_font.color=}")

            # cell.font = cell.font.copy(**cell_font_args)
            # cell.font = openpyxl.styles.Font(**cell.font.__dict__, **cell_font_args)
            cell.font = new_cell_font
    else:
        logger.error("feature not supported on xls file - only XLSX")
        raise NotImplementedError(
            "feature not supported on xls file - only XLSX"
        )


def getExcelCellPatternFill(
        excel_dict: dict,
        row: int,
        col_name: str,
        debug: bool = False
) -> tuple[str | None, str | None, str | None, str | None, str | None]:
    """
    For a defined row number and a column name string, get the cell_font information for that cell

    Inputs:
        excel_dict - custom excel dictionary created to represent a worksheet object
        row - int - row number that we are extracting from
        col_name - str - the header name associated with the column
        debug - bool - when true, we display debugging messages

    Returns:
        cell_color
        cell_fill_type
        cell_start_color
        cell_end_color
        cell_fill - the PatternFill object of this cell

    """

    cell_color = None
    cell_fill_type = None
    cell_start_color = None
    cell_end_color = None

    if debug:
        print("setExcelCellPatternFill:excel_dict:", excel_dict)
        print("setExcelCellPatternFill:row:", row)
        print("setExcelCellPatternFill:col_name:", col_name)
    logger.debug("excel_dict:%s", excel_dict)
    logger.debug("row:%s", row)
    logger.debug("col_name:%s", col_name)

    # determine the col # we are using by doing a header lookup for col_name
    # add to that the offset if the starting column is not the first columnn
    # if the col_name provided was not in the header - it will automatically evoke a ValueError reporting the problem
    # so we will just allow that to be the error message
    col = excel_dict["header"].index(col_name) + excel_dict["sheetmincol"]

    # debugging
    if debug:
        print("pattern")
        print("col_name:", col_name)
        print("col:", col)
        print("row:", row)
        print(
            "value:",
            excel_dict["s"]
            .cell(row=row + 1 + excel_dict["row_header"], column=col + 1)
            .value,
        )

    # return none if no style
    if (
        not excel_dict["s"]
        .cell(row=row + 1 + excel_dict["row_header"], column=col + 1)
        .has_style
    ):
        return (
            cell_color,
            cell_fill_type,
            cell_start_color,
            cell_end_color,
            None,
        )

    # get cell value
    if excel_dict["xlsxfiletype"]:
        # get fill settings
        cell_fill = (
            excel_dict["s"]
            .cell(row=row + 1 + excel_dict["row_header"], column=col + 1)
            .fill
        )
        # debugging
        if debug:
            print(f"{cell_fill=}\n{'-' * 40}")
            print("setExcelCellPatternFill:fill_type:", cell_fill.fill_type)
            for x in [
                y for y in dir(cell_fill) if "__" not in y and y != "copy"
            ]:
                print(
                    "setExcelCellPatternFill:" + str(x) + ":",
                    getattr(cell_fill, x),
                )

        # special processing based on type
        if cell_fill.fill_type in ("solid", "darkHorizontal"):
            cell_color = cell_fill.fgColor.rgb
            cell_fill_type = cell_fill.fill_type
            cell_start_color = cell_fill.start_color
            cell_end_color = cell_fill.end_color
        else:
            cell_fill_type = cell_fill.fill_type

        # return cell_fill
        return (
            cell_color,
            cell_fill_type,
            cell_start_color,
            cell_end_color,
            cell_fill,
        )

    else:
        logger.error("feature not supported on xls file - only XLSX")
        raise NotImplementedError(
            "feature not supported on xls file - only XLSX"
        )


def setExcelCellPatternFill(
    excel_dict: dict,
    row: int,
    col_name: str,
    fill: str | None = None,
    start_color: str | None = None,
    end_color: str | None = None,
    fg_color: str | None = None,
    fill_type: str = "solid",
    debug: bool = False,
) -> None:
    """
    For a defined row number and a column name string, set the PatternFill attributes for a cell in excel

    If you wnat to clear that setting - you must pass in the string "None" not the None value

    Inputs:
        excel_dict - custom excel dictionary created to represent a worksheet object
        row - the row in the data
        col_name - the name of the column we are setting

    Optional Inputs:
        fill - PatternFill object
        fg_color - specify the color of hte fill - if it is just one color and not one that changes from start to end
        start_color - Specify the color of the fill using hex color codes.
        end_color - Specify the color of the fill using hex color codes.
        fill_type - Specify the type of fill. Common types include:
            solid: Solid color fill.
            gray125: Light gray fill.
            lightDown: Light diagonal stripes.
            lightUp: Light diagonal stripes in the opposite direction.
            darkDown: Dark diagonal stripes.
            darkUp: Dark diagonal stripes in the opposite direction.

        to clear these values pass in "None" string that will cause the optoin to be set to the None variable
        if None value is passed in we skip updating that attribute

    """

    # test inputs
    if (start_color and start_color != "None") and not (
        end_color and end_color != "None"
    ):
        raise ValueError(
            "start_color populated and end_color not - both must be populated - or just set fg_color"
        )
    if (end_color and end_color != "None") and not (
        start_color and start_color != "None"
    ):
        raise ValueError(
            "end_color populated and start_color not - both must be populated - or just set fg_color"
        )
    if (fg_color and fg_color != "None") and any(
        [
            (start_color and start_color != "None"),
            (end_color and end_color != "None"),
        ]
    ):
        raise ValueError(
            "fg_color populated and either start_color or end_color is populated - please clear start_color and/or end_color"
        )

    # make sure fill is set properly
    if (
        fill
        and fill != "None"
        and not isinstance(fill, openpyxl.styles.fills.PatternFill)
    ):
        raise TypeError(f"fill must be PatternFile type but is: {type(fill)}")

    # override fill - if we want to clear a value we must pass it in as 'None'
    override_fill = {}
    if fill is not None:
        override_fill["fill"] = None if fill == "None" else fill
    if fill_type is not None:
        override_fill["fill_type"] = None if fill_type == "None" else fill_type
    if start_color is not None:
        override_fill["start_color"] = (
            None if start_color == "None" else start_color
        )
    if end_color is not None:
        override_fill["end_color"] = None if end_color == "None" else end_color
    if fg_color is not None:
        override_fill["fgColor"] = None if fg_color == "None" else fg_color

    if debug:
        print("setExcelCellPatternFill:excel_dict:", excel_dict)
        print("setExcelCellPatternFill:row:", row)
        print("setExcelCellPatternFill:col_name:", col_name)
        print("setExcelCellPatternFill:fill-type:", type(fill))

    # determine the col # we are using by doing a header lookup for col_name
    # add to that the offset if the starting column is not the first columnn
    # if the col_name provided was not in the header - it will automatically evoke a ValueError reporting the problem
    # so we will just allow that to be the error message
    col = excel_dict["header"].index(col_name) + excel_dict["sheetmincol"]

    # set cell Pattern value
    if excel_dict["xlsxfiletype"]:
        # get fill settings
        cell_fill = (
            excel_dict["s"]
            .cell(row=row + 1 + excel_dict["row_header"], column=col + 1)
            .fill
        )

        if fill:
            # passed in the fill type object - set it
            excel_dict["s"].cell(
                row=row + 1 + excel_dict["row_header"], column=col + 1
            ).fill = fill
            if debug:
                print("Pattern set by fill")
        elif start_color:
            # if we updated start_color and end_color
            # now update this fill with the values passed in
            excel_dict["s"].cell(
                row=row + 1 + excel_dict["row_header"], column=col + 1
            ).fill = openpyxl.styles.PatternFill(
                fill_type=override_fill.get("fill_type", cell_fill.fill_type),
                start_color=override_fill.get(
                    "start_color", cell_fill.start_color
                ),
                end_color=override_fill.get("end_color", cell_fill.end_color),
            )
            if debug:
                print("Pattern set by start_color/end_color")
        else:
            # if we updated fgColor
            # now update this fill with the values passed in
            excel_dict["s"].cell(
                row=row + 1 + excel_dict["row_header"], column=col + 1
            ).fill = openpyxl.styles.PatternFill(
                fill_type=override_fill.get("fill_type", cell_fill.fill_type),
                fgColor=override_fill.get("fgColor", cell_fill.fgColor),
            )
            if debug:
                print("Pattern set by fgColor")
    else:
        logger.error("feature not supported on xls file - only XLSX")
        raise NotImplementedError(
            "feature not supported on xls file - only XLSX"
        )


# copy the cell formatting from src into out cell by cell - this is color and fill
def copyExcelCellFmtOnRow(
    excel_dict_src: dict,
    src_row: int,
    excel_dict_out: dict,
    out_row: int,
    debug: bool = False,
) -> None:
    """
    Read in the format defintion for a worksheet (src), and input row
    and copy cell by cell for each column based on the column name in both worksheets
    to an output worksheet (out)

    Inputs:
        excel_dict_src - custom excel dictionary created to represent a worksheet object that is the source for formatting
        src_row - int - row number that we are extracting source formatting from
        excel_dict_out - custom excel dictionary created to represent a worksheet object that is where we assign formatting
        out_row - int - row number that we are placing formatting into
        debug - bool - when true, we display debugging messages

    """

    # step through the output columns - find the input column with the sname name and copy over the formatting
    for col_name in excel_dict_out["header"]:
        # validate the out column exists in the source
        if col_name not in excel_dict_src["header"]:
            # not in the input worksheet so skip
            continue

        # grab the color and field for this row/column
        fg_color, fill_type, start_color, end_color, cell_fill = (
            getExcelCellPatternFill(
                excel_dict_src, src_row, col_name, debug=debug
            )
        )

        # debugging
        if debug:
            print(
                f"{src_row=}, {fg_color=}, {fill_type=}, {start_color=}, {end_color=}"
            )

        # take no action if if there is no format to copy over
        if (
            fill_type is None
            and fg_color is None
            and start_color is None
            and end_color is None
        ):
            continue

        # force to fg_color if set - need to understand this better
        if fg_color:
            start_color = None
            end_color = None
            if debug:
                print("fg_color is set, clearing start_color and end_color")

        # now copy this over to the out worksheet
        setExcelCellPatternFill(
            excel_dict_out,
            out_row,
            col_name,
            fill=None,
            start_color=start_color,
            end_color=end_color,
            fg_color=fg_color,
            fill_type=fill_type,
            debug=debug,
        )


def setExcelColumnValue(
    excel_dict: dict, col_name: str, value: Any = "", debug: bool = False
) -> None:
    """
    Set the value of a column, defined by col_name, to the value passed in for all cells in range

    Find the column, then set all cell values in that column
    Then iterate through that column and set the values

    Inputs:
        excel_dict - custom excel dictionary created to represent a worksheet object
        col_name - str - the header name associated with the column
        value - Any - the value to put in the cells in thsi column
        debug - bool - when true, we display debugging messages


    """

    # step through all rows for a defined column name and set the value to the single value passed in
    for row in range(excel_dict["row_header"] + 1, excel_dict["sheetmaxrow"]):
        setExcelCellValue(excel_dict, row, col_name, value, debug)


def any_field_is_populated(
    rec: dict, fldlist: list[str], debug: bool = False
) -> bool:
    """
    Return a TRUE if any of the 'fldlist' elements in rec is populated
    the field has a value or the field is not a string

    Inputs:
        rec - dict - the record being evaluated
        fldlist - list - list of fields to check to see if they are populated
        debug - bool - when true - dsiplay run time messages

    Returns
        is_populated - bool - true when any of the values in fldlist is populated

    """
    # test inputs
    if not rec:
        return False
    if not isinstance(rec, dict):
        raise TypeError(f"rec must be dict but is: {type(rec)}")
    if not fldlist:
        raise ValueError("fldlist must be populated")
    if not isinstance(fldlist, list):
        raise TypeError(f"fldlist must be list but is: {type(fields)}")

    # validate fldlist
    for fld in fldlist:
        # current conditions - if it returns true or has a length
        if rec[fld]:
            if debug:
                print(f"fld [{fld}] is popuated it has a value")
            return True
        elif not isinstance(rec[fld], str):
            if debug:
                print(f"fld [{fld}] is populated because it is not a string")
            return True
        else:
            if debug:
                print(f"fld [{fld}] is not populated")

    if debug:
        print("no fldlist are populated - returning false")

    return False


def create_multi_key_lookup_excel(
    excel_dict: dict,
    fldlist: list[Any],
    copy_fields: list[Any] | None = None,
    debug: bool = False,
) -> dict:
    """
    Create a multi key dictionary that gets to the record based on the
    keys in the record

    if user sets the copy_fields with the list of fields that can have values
    then we check the record
    to determine if any of the fields has a value, and if none have a value we skip
    that record

    Inputs:
        excel_dict - custom excel dictionary created to represent a worksheet object
        fldlist - list - the list of column names that make the unique bueinsss key
        copy_fields - list|None - defines the list of fields that we would want to copy over when populated
        debug - bool - when true, we display debugging messages

    Returns
        src_lookup - dict - a multi key dictionary with the final value being the record itself.

    """

    # set value if not set
    if copy_fields is None:
        copy_fields = list()

    # test inputs
    if not isinstance(excel_dict, dict):
        raise TypeError(f"excel_dict must be type dict but is: {type(fldlist)}")
    if not isinstance(fldlist, list):
        raise TypeError(f"fldlist must be type list but is: {type(fldlist)}")

    # check that the copy_fields keys are in the first record
    if copy_fields and not isinstance(copy_fields, list):
        raise TypeError(
            f"copy_fields must be type - list - but is: {type(copy_fields)}"
        )

    # validate keys passed in match column names we found
    bad_keys = [x for x in fldlist if x not in excel_dict["header"]]
    if bad_keys:
        raise ValueError(
            f"fldlist values not in the header: {','.join(bad_keys)}"
        )
    bad_keys = [x for x in copy_fields if x not in excel_dict["header"]]
    if bad_keys:
        raise ValueError(
            f"copy_fields values not in the header: {','.join(bad_keys)}"
        )
    #
    # set up the dictionary to be populated
    src_lookup = {}

    # step through each record
    for row in range(excel_dict["row_header"] + 1, excel_dict["sheetmaxrow"]):
        # test that this record has values in the copy_fields attributes
        ## TODO - build out this logic
        # if False and copy_fields and not any_field_is_populated(row, copy_fields):
        # no values set in copy_fields has a value so we don't convert this record
        # continue

        # get the first key and value
        fld = fldlist[0]
        fldvalue = getExcelCellValue(excel_dict, row, fld)
        if fldvalue not in src_lookup:
            if len(fldlist) > 1:
                # multi key
                src_lookup[fldvalue] = {}
            else:
                # single key - set the value
                src_lookup[fldvalue] = row

        # now create the changing key
        ptr = src_lookup[fldvalue]

        # get all other key/value after the first
        for fld in fldlist[1:]:
            # get the value
            fldvalue = getExcelCellValue(excel_dict, row, fld)
            # check to see this level is working
            if fldvalue not in ptr:
                ptr[fldvalue] = {}
            # if we are on the last fld then set to rec
            if fld == fldlist[-1]:
                ptr[fldvalue] = row
            else:
                # update the ptr
                ptr = ptr[fldvalue]

    return src_lookup


# ----------------------------------------


def calc_col_mapping(rec: dict) -> tuple[str, dict]:
    """
    take a dict that is a record that has FLD_XLSCOL_ABS as one of the keys
    done by loaded with "save_col_abs" flag in option dict set to true
    and create a column header to column number mapping
    that is then converted to json

    Inputs:
        rec - dict - a record that has been read in with FLD_XLSCOL_ABS as one of the attributes

    Returns:
        json string - str - string version of col_mapping
        col_mapping - dict -

    Primative

    """

    # check to see if the needed field is there
    if FLD_XLSCOL_ABS not in rec:
        raise ValueError(
            f"[{FLD_XLSCOL_ABS}] not in record - read file with save_col_abs flag enabled in optiondict"
        )

    # step thorugh each of the keys in the record and build up a dictionary that defines the column
    # column number that each header would be in
    col_mapping = {
        fld: rec[FLD_XLSCOL_ABS] + idx
        for idx, fld in enumerate(list(rec.keys()))
    }

    return json.dumps(col_mapping), col_mapping


def set_col_mapping(rec: dict) -> None:
    """
    calculate and add column mapping to a single record into the FLD_XLSNEW_COLMAP fields in the record
    this is an inplace update where we add a new key to the rec dictionary

    Inputs:
        rec - dict - the record read in from xlsx

    Returns:
        Nothing

    Primative

    """

    # get the mapping defined
    col_mapping_str, col_mapping = calc_col_mapping(rec)
    # assign to this one record
    rec[FLD_XLSNEW_COLMAP] = col_mapping_str


def set_col_mapping_list(records: list[dict]) -> None:
    """
    calcuate the column mapping for this list or records that were read in
    and add column mapping to all records in thie list

    this routine is generally called when flags are set to add this value to read in data

    Inputs:
        records - list[dict] - list of records read in

    Returns:
        Nothing

    """

    # get the mapping defined
    col_mapping_str, col_mapping = calc_col_mapping(records[0])
    # now assign to all records
    for x in records:
        # set this across all records
        x[FLD_XLSNEW_COLMAP] = col_mapping_str


def extract_col_mapping(rec: dict) -> tuple[dict, str]:
    """
    take a dict that is a record that has FLD_BOM_NEW_COLMAP as one of the keys and
    extract the column header to column number mapping
    and return string and dict version of this

    Inputs:
        rec - dict - the record read in from xlsx

    Returns:
        col_mapping - dict
        col_mapping_str - str

    Primative

    """
    # test for existance
    if FLD_XLSNEW_COLMAP not in rec:
        raise ValueError(
            "Column mapping column not in record: " + FLD_XLSNEW_COLMAP
        )

    # get the column mapping out of the record
    col_mapping_str = rec[FLD_XLSNEW_COLMAP]
    col_mapping = json.loads(col_mapping_str)

    return col_mapping, col_mapping_str


# -------- READ FILES -------------------------


def readxls2list(
    xlsfile: str | os.PathLike,
    sheetname: str | None = None,
    save_row: bool = False,
    optiondict: dict | None = None,
    debug: bool = False,
) -> list[dict]:
    """
    Read in a XLS or XLSX file
    read in the XLS and create a dictionary to the records
    assumes the first line of the XLS file is the header/defintion of the XLS

    Inputs:
        xlsfile - the filename/path to the file being read in
        sheetname - if populated, the worksheet name in the excel file we should get data from
        save_row - bool - when enabled, we add a column to each record that defines the row(XLSRow) in the XLS where the line came from
        optiondict - dict - key/value pairs that control behaviors
        debug - bool - when enabled, we display processing messages to track and debug what is going on.

    Returns
        data - list of dictionaries - the data extarcted from the xls/worksheet as a list of dictionaries

    """
    if optiondict is None:
        optiondict = {"col_header": True, "save_row": save_row}
    else:
        optiondict["col_header"] = True
        optiondict["save_row"] = save_row
    # set the option if it is populated
    if sheetname:
        optiondict["sheetname"] = sheetname
    return readxls2list_findheader(
        xlsfile, [], optiondict=optiondict, debug=debug
    )


def readxls2dict(
    xlsfile: str,
    dictkeys: list[str],
    sheetname: str | None = None,
    save_row: bool = False,
    dupkeyfail: bool = False,
    optiondict: dict | None = None,
    debug: bool = False,
) -> dict:
    """
    read in the XLS and create a dictionary to the records
    based on one or more key fields
    assumes the first line of the CSV file is the header/defintion of the CSV


    """
    if optiondict is None:
        optiondict = {"col_header": True, "save_row": save_row}
    else:
        optiondict["col_header"] = True
        optiondict["save_row"] = save_row
    # set sheetname if populated
    if sheetname:
        optiondict["sheetname"] = sheetname
    return readxls2dict_findheader(
        xlsfile,
        dictkeys,
        [],
        optiondict=optiondict,
        debug=debug,
        dupkeyfail=dupkeyfail,
    )


def readxls2dump(
    xlsfile: str,
    rows: int = 10,
    sep: str = ":",
    no_warnings: bool = False,
    returnrecs: bool = False,
    sheet_name_col: str | None = None,
    debug: bool = False,
) -> list:
    """
    read in the xls - output the first XX lines

    """

    if sheet_name_col is None:
        sheet_name_col = "sheet_name"

    fmtstr1 = sep.join(("{}", "{}", "{}", "{}", "{}")) + sep
    fmtstr2 = sep.join(("{}", "{}", "{:02d}", "{:03d}", "{}")) + sep
    recheader = ["xlsfile", sheet_name_col, "reccnt", "colcnt", "value"]
    xlslines = []
    xlsrecs = []
    optiondict = {
        "no_header": True,
        "aref_result": True,
        "save_row": True,
        "max_rows": rows + 5,
        "no_warnings": no_warnings,
    }
    excel_dict = readxls_findheader(
        xlsfile, [], optiondict=optiondict, debug=debug
    )
    xlslines.append(fmtstr1.format(*recheader))
    for sheetname in excel_dict["sheet_names"]:
        if debug:
            print(sheetname, "-" * 80)
        optiondict["sheetname"] = sheetname
        excel_dict = chgsheet_findheader(
            excel_dict, [], optiondict=optiondict, debug=debug
        )
        results = excelDict2list_findheader(
            excel_dict, [], optiondict=optiondict, debug=debug
        )
        reccnt = 0
        for rec in results:
            colcnt = 0
            for col in rec:
                xlslines.append(
                    fmtstr2.format(
                        excel_dict["xlsfile"],
                        excel_dict["sheet_name"],
                        reccnt,
                        colcnt,
                        col,
                    )
                )
                if returnrecs:
                    xlsrecs.append(
                        dict(
                            zip(
                                recheader,
                                [
                                    excel_dict["xlsfile"],
                                    excel_dict["sheet_name"],
                                    reccnt,
                                    colcnt,
                                    col,
                                ],
                            )
                        )
                    )
                colcnt += 1
            reccnt += 1
            if reccnt > rows:
                break
    if returnrecs:
        return xlslines, xlsrecs
    else:
        return xlslines


def readxls2list_all_sheets(
    xlsfile: str,
    req_cols: list[str],
    xlatdict: dict | None = None,
    optiondict: dict | None = None,
    col_aref: list[str] | None = None,
    data_only: bool = True,
    debug: bool = False,
) -> Tuple[dict | None, list[str] | None, dict | None]:
    """
    This routine opens the xlsx and reads teh data in from all sheets -
    but teh column headers have to be same across sheets

    read in workbook with multiple sheets - hopefully each sheet is the same structrue
    and pull out all data from them - finding the header and then getting a list of dicts
    return a dictionary keyed by sheetname, with values of the list of dicts that made up the data in that sheet

    Inputs:
        xlsfile - str - filename/path to the excel file being read
        req_cols - list - column names we want as the outputs for in each sheet
        xlatdict - dict - mapping from what may exist in sheet and how it maps to a value in req_cols
        col_aref - list
        data_only - bool
        debug - bool - when true, we display processing messages used in debugging.

    Returns:
        a dict where the key is the sheet name and values list of dicts read in for that sheet
        a list of the headers captured in the right order for the first sheet read in
        a dict by sheetname of errors found when reading in sheets

    """

    if optiondict is None:
        optiondict = dict()
    if xlatdict is None:
        xlatdict = dict()
    if col_aref is None:
        col_aref = list()

    # test inputs
    if not req_cols:
        raise ValueError(
            "req_cols must be populated with the list of required column headers"
        )
    if not isinstance(req_cols, list):
        raise TypeError(f"req_cols must be type list but is: {type(req_cols)}")
    if not isinstance(col_aref, list):
        raise TypeError(f"col_aref must be type list but is: {type(col_aref)}")

    # capture the passed in sheet name and make sure we return it
    origsheetname = None
    if "sheetname" in optiondict:
        origsheetname = optiondict["sheetname"]

    # first open the xlsx file
    excel_dict = readxls_excelDict(
        xlsfile,
        req_cols,
        xlatdict=xlatdict,
        optiondict=optiondict,
        col_aref=col_aref,
        data_only=data_only,
        debug=debug,
    )

    if debug:
        print("excel_dict")
        pprint.pprint(excel_dict)

    # return if we got nothing
    if excel_dict is None:
        return excel_dict, None, None

    # capture the first header
    header_first_sheet = excel_dict["header"]

    # create a dict with the results of each return value
    all_sheet_data = {}
    sheet_error = {}

    #  DEBUGGING - read in the data from this sheet - first time
    if False:
        all_sheet_data[excel_dict["sheet_name"]] = excelDict2list_findheader(
            excel_dict,
            req_cols,
            xlatdict=xlatdict,
            optiondict=optiondict,
            col_aref=col_aref,
            debug=debug,
        )

        return all_sheet_data, None, None

    # step through each sheetname and get the list of recrods for that sheet
    for s in excel_dict["sheet_names"]:
        # change to this sheet
        optiondict["sheetname"] = s

        # attempt to do this
        try:
            if debug:
                print("change sheetname to: ", s)
            # change the the sheet we are interested
            excel_dict = chgsheet_findheader(
                excel_dict,
                req_cols,
                xlatdict=xlatdict,
                optiondict=optiondict,
                col_aref=col_aref,
                data_only=data_only,
                debug=debug,
            )
        except Exception as e:
            if debug:
                print("Except - change sheet - sheet: ", s, "Error: ", e)
            sheet_error[s] = str(e)
            all_sheet_data[s] = []
            continue

        # set the header if not set
        if not header_first_sheet:
            header_first_sheet = excel_dict["header"]

        # override this if we passed in a sheetname and this is the sheetname
        if origsheetname and s == origsheetname:
            header_first_sheet = excel_dict["header"]

        # extract the data
        all_sheet_data[s] = excelDict2list_findheader(
            excel_dict,
            req_cols,
            xlatdict=xlatdict,
            optiondict=optiondict,
            col_aref=col_aref,
            debug=debug,
        )

        if debug:
            print("records from sheet:", s, len(all_sheet_data[s]))
            print("showing all_sheet_data key and count")
            for k, v in all_sheet_data.items():
                print(k, len(v))
            print("-" * 40)

    # return the value to its original value
    if origsheetname is None:
        del optiondict["sheetname"]
    else:
        optiondict["sheetname"] = origsheetname

    if debug:
        print("end of routine - show what we collected per sheet:")
        for k, v in all_sheet_data.items():
            print(k, len(v))

    return all_sheet_data, header_first_sheet, sheet_error


# ---------- GENERIC OPEN EXCEL to enable EDIT ----------------------
#
# or passed on to other routines to extract the data for processing
#
# Open to edit and save:
# # example how to use:  open file for editting
# xls = kvxls.readxls_findheader( 'Wine Collection 20-05-07-v02.xlsm', [],
# optiondict={'col_header' : True}, data_only=False )
#
# # change a cell
# kvxls.setExcelCellValue( xls, 2, 'Rating', 'Changed')
# # save the file
# kvxls.writexls( xls, 'newfile.xlsm' )
#
#
# generic routine that reads in the XLS and returns back a dictionary for that xls
# that is either used to interact with that XLS object, or is passed to other routines
# that then create the dictionary/list of that xls and then close out that XLS.
#    data_only - when set to FALSE - will allow you to read macro enable file and update directly
#                and save the updated file
def readxls_excelDict(
    xlsfile: str,
    req_cols: list[str] | None = None,
    xlatdict: dict | None = None,
    optiondict: dict | None = None,
    col_aref: list[str] | None = None,
    data_only: bool = True,
    debug: bool = False,
) -> dict:
    """
    Open the excel file and prepare for doing other work, create an excel object to be used for future work

    """

    # set values if they were not set
    if xlatdict is None:
        xlatdict = dict()
    if optiondict is None:
        optiondict = dict()
    if col_aref is None:
        col_aref = list()

    # type check
    if not isinstance(col_aref, list):
        raise TypeError(f"col_aref must be type list but is: {type(col_aref)}")
    if not isinstance(req_cols, list):
        raise TypeError(f"req_cols must be type list but is: {type(req_cols)}")
    if not isinstance(optiondict, dict):
        raise TypeError(
            f"optiondict must be type dict but is: {type(optiondict)}"
        )
    if not isinstance(xlatdict, dict):
        raise TypeError(f"xlatdict must be type dict but is: {type(xlatdict)}")

    # create the excel configuratoin
    excel_config = create_excel_config(
        optiondict=optiondict, func_name="kvxls.readxls_excelDict"
    )

    # add passed in values
    excel_config.req_cols = req_cols
    excel_config.col_aref = col_aref
    excel_config.xlatdict = xlatdict
    excel_config.data_only = data_only

    # local variables - not used so commented out
    # header = None

    # debugging
    if debug:
        print("req_cols:", req_cols)
        print("xlatdict:", xlatdict)
        print("optiondict:", optiondict)
        print("col_aref:", col_aref)
        print("excel_config:", excel_config)

    # set up variables
    start_row = 0  # if passed in - we start the search at this row (starts at 1 or greater)
    max_rows = 100000000
    keep_vba = False

    if debug:
        print("optiondict after badoption_check:", optiondict)

    if "start_row" in optiondict:
        start_row = (
            optiondict["start_row"] - 1
        )  # because we are not ZERO based in the users mind

    if "max_rows" in optiondict:
        max_rows = optiondict["max_rows"]
    if "keep_vba" in optiondict:
        keep_vba = optiondict["keep_vba"]

    # debugging
    if debug:
        print("readxls_excelDict")
        print("req_cols:", req_cols)
        print("col_aref:", col_aref)
        print("start_row:", start_row)
        print("optiondict:", optiondict)
        print("excel_config:", excel_config)

    # determine what filetype we have here
    xlsxfiletype = xlsfile.endswith(".xlsx") or xlsfile.endswith(".xlsm")
    excel_config.xlsxfiletype = xlsxfiletype

    logger.debug("readxls_excelDict:excel_config: %s", excel_config)

    # debugging
    logger.debug("xlsxfiletype:%s", xlsxfiletype)

    # Load in the workbook (set the data_only=True flag to get the value on the formula)
    if xlsxfiletype:
        # XLSX file
        if data_only:
            wb = openpyxl.load_workbook(xlsfile, data_only=True)
        else:
            wb = openpyxl.load_workbook(
                xlsfile, read_only=False, keep_vba=keep_vba
            )
        sheet_names = wb.sheetnames
    elif XLSXONLY:
        # put in to deal with a simplified library
        raise NotImplementedError(
            "this library only supports newer Excel file types"
        )
    else:
        # XLS file
        wb = xlrd.open_workbook(xlsfile)
        sheet_names = wb.sheet_names()

    # debugging
    if debug:
        print("sheet_names:", sheet_names)

    # get the sheet we are going to work with
    if "sheetname" in optiondict and optiondict["sheetname"]:
        # user defined by name
        sheet_name = optiondict["sheetname"]
    elif "sheetrow" in optiondict:
        # user defined by index in the list of names
        sheet_name = sheet_names[optiondict["sheetrow"]]
    else:
        # take the first name
        sheet_name = sheet_names[0]

    # debugging
    if debug:
        print("sheet_name:", sheet_name)

    # create a workbook sheet object - using the name to get to the right sheet
    if xlsxfiletype:
        s = wb[sheet_name]
        sheettitle = s.title
        sheetmaxrow = s.max_row
        sheetmaxcol = s.max_column
        sheetminrow = 0
        sheetmincol = 0
    else:
        s = wb.sheet_by_name(sheet_name)
        sheettitle = s.name
        sheetmaxrow = s.nrows
        sheetmaxcol = s.ncols
        sheetminrow = 0
        sheetmincol = 0

    # debugging
    if debug:
        print("sheettitle:", sheettitle)
        print("sheetmaxrow:", sheetmaxrow)
        print("sheetmaxcol:", sheetmaxcol)

    # check and see if we need to limit max row
    if max_rows < sheetmaxrow:
        sheetmaxrow = max_rows
        if debug:
            print("sheetmaxrow-changed:", sheetmaxrow)

    # ------------------------------- HEADER START ------------------------------

    # ------------------------------- HEADER END ------------------------------

    # ------------------------------- OBJECT DEFINITION ------------------------------
    excel_dict = {
        "xlsfile": xlsfile,
        "xlsxfiletype": xlsxfiletype,
        "keep_vba": keep_vba,
        "wb": wb,
        "sheet_names": sheet_names,
        "sheet_name": None,
        "s": None,
        "sheettitle": sheettitle,
        "sheetmaxrow": sheetmaxrow,
        "sheetmaxcol": sheetmaxcol,
        "sheetminrow": sheetminrow,
        "sheetmincol": sheetmincol,
        "row_header": None,
        "header": None,
        "start_row": None,
    }

    if debug:
        print("excel_dict: ", excel_dict)

    return excel_dict


def readxls_findheader(
    xlsfile: str,
    req_cols: list[str],
    xlatdict: dict | None = None,
    optiondict: dict | None = None,
    col_aref: list[str] | None = None,
    data_only: bool = True,
    debug: bool = False,
) -> dict:
    """
    generic routine that reads in the XLS and returns back a dictionary for that xls
    that is either used to interact with that XLS object, or is passed to other routines
    that then create the dictionary/list of that xls and then close out that XLS.
         data_only - when set to FALSE - will allow you to read macro enable file and update directly
                     and save the updated file
    """

    # defaults for not set values
    if xlatdict is None:
        xlatdict = {}
    if optiondict is None:
        optiondict = {}

    # type check
    if col_aref is not None and not isinstance(col_aref, list):
        raise TypeError(f"col_aref must be type list but is: {type(col_aref)}")
    if not isinstance(req_cols, list):
        raise TypeError(f"req_cols must be type list but is: {type(req_cols)}")
    if not isinstance(optiondict, dict):
        raise TypeError(
            f"optiondict must be type dict but is: {type(optiondict)}"
        )
    if not isinstance(xlatdict, dict):
        raise TypeError(f"xlatdict must be type dict but is: {type(xlatdict)}")

    # create the excel configuratoin
    excel_config = create_excel_config(
        optiondict=optiondict, func_name="kvxls.readxls_findheader"
    )

    # add passed in values
    excel_config.filename = xlsfile
    excel_config.req_cols = req_cols
    excel_config.col_aref = col_aref
    excel_config.xlatdict = xlatdict

    # determine what filetype we have here
    xlsxfiletype = xlsfile.endswith(".xlsx") or xlsfile.endswith(".xlsm")
    excel_config.xlsxfiletype = xlsxfiletype

    # debugging
    if debug:
        print("req_cols:", req_cols)
        print("xlatdict:", xlatdict)
        print("optiondict:", optiondict)
        print("col_aref:", col_aref)
        print("excel_config:", excel_config)

    # local variables
    header = None

    # set flags
    col_header = (
        False  # if true - we take the first row of the file as the header
    )
    no_header = False  # if true - there are no headers read - we either return
    aref_result = False  # if true - we don't return dicts, we return a list
    save_row = False  # if true - then we append/save the XLSRow with the record
    save_row_abs = False  # if true - then we append/save the openpxl row
    save_col_abs = False  # if true - then we append/save the absolute xlsx column number of the first column - from openpyxl
    save_colmap = False  # if true - then we add a new field that housed the colmapp for this
    save_colfmt = None  # if populated with a column header - than we capture and save the format of that column
    keep_vba = True  # if true - then load the xlsx with vba scripts on and save as xlsm
    allow_empty = (
        False  # if true - we allow a header to be read in with no data
    )
    row_header = None  # we will set this later

    start_row = 0  # if passed in - we start the search at this row (starts at 1 or greater)

    max_rows = 100000000

    if debug:
        print("optiondict after badoption_check:", optiondict)

    # pull in passed values from optiondict
    if "col_header" in optiondict:
        col_header = optiondict["col_header"]
    if "aref_result" in optiondict:
        aref_result = optiondict["aref_result"]
    if "no_header" in optiondict:
        no_header = optiondict["no_header"]
    if "allow_empty" in optiondict:
        allow_empty = optiondict["allow_empty"]
    if "start_row" in optiondict:
        start_row = (
            optiondict["start_row"] - 1
        )  # because we are not ZERO based in the users mind
    if "save_row" in optiondict:
        save_row = optiondict["save_row"]
    if "save_row_abs" in optiondict:
        save_row_abs = optiondict["save_row_abs"]
    if "save_col_abs" in optiondict:
        save_col_abs = optiondict["save_col_abs"]
    if "save_colmap" in optiondict:
        save_colmap = optiondict["save_colmap"]
    if "save_colfmt" in optiondict:
        save_colfmt = optiondict["save_colfmt"]
    if "max_rows" in optiondict:
        max_rows = optiondict["max_rows"]
    if "keep_vba" in optiondict:
        keep_vba = optiondict["keep_vba"]

    # debugging
    if debug:
        print("readxls_findheader")
        print("req_cols:", req_cols)
        print("col_aref:", col_aref)
        print("col_header:", col_header)
        print("aref_result:", aref_result)
        print("no_header:", no_header)
        print("start_row:", start_row)
        print("save_row:", save_row)
        print("save_row_abs:", save_row_abs)
        print("save_col_abs:", save_col_abs)
        print("save_colmap:", save_colmap)
        print("save_colfmt:", save_colfmt)
        print("allow_empty:", allow_empty)
        print("optiondict:", optiondict)
    logger.debug("req_cols:%s", req_cols)
    logger.debug("col_aref%s", col_aref)
    logger.debug("col_header:%s", col_header)
    logger.debug("aref_result:%s", aref_result)
    logger.debug("no_header:%s", no_header)
    logger.debug("start_row:%s", start_row)
    logger.debug("save_row:%s", save_row)
    logger.debug("save_row_abs:%s", save_row_abs)
    logger.debug("save_col_abs:%s", save_col_abs)
    logger.debug("save_colmap:%s", save_colmap)
    logger.debug("save_colfmt:%s", save_colfmt)
    logger.debug("allow_empty:%s", allow_empty)
    logger.debug("optiondict:%s", optiondict)

    # special condidtiaon for no header
    if not col_header and not req_cols and start_row and col_aref:
        no_header = True
        excel_config.no_header = True
        if debug:
            print(
                "Setting no_header because of col_header, start_row, col_aref"
            )
            print("no_header:", no_header)
    elif debug:
        print(col_header, start_row, col_aref)

    # another special condition
    if no_header and not col_aref and not aref_result:
        # there is no header in the file
        # we did not pass in a header
        # and they did not set this as aref_results
        # we must force it to aref_result
        aref_result = True
        excel_config.aref_result = True

        if debug:
            print(
                "no_header true, col_aref not populated and aref_result False - force aref_result to True"
            )

    # build object that will be used for record matching
    p = kvmatch.MatchRow(
        req_cols,
        xlatdict,
        optiondict,
        optiondict2={"noshowwarning": True, "fix_missing": True},
        debug=debug,
    )

    # determine what filetype we have here
    xlsxfiletype = xlsfile.endswith(".xlsx") or xlsfile.endswith(".xlsm")

    # debugging
    logger.debug("xlsxfiletype:%s", xlsxfiletype)

    # Load in the workbook (set the data_only=True flag to get the value on the formula)
    if xlsxfiletype:
        # XLSX file
        if data_only:
            wb = openpyxl.load_workbook(xlsfile, data_only=True)
        else:
            wb = openpyxl.load_workbook(
                xlsfile, read_only=False, keep_vba=keep_vba
            )
        sheet_names = wb.sheetnames
    elif XLSXONLY:
        # put in to deal with a simplified library
        raise NotImplementedError(
            "this library only supports newer Excel file types"
        )
    else:
        # XLS file
        wb = xlrd.open_workbook(xlsfile)
        sheet_names = wb.sheet_names()

    # debugging
    if debug:
        print("sheet_names:", sheet_names)
    logger.debug("sheet_names:%s", sheet_names)

    # get the sheet we are going to work with
    if "sheetname" in optiondict and optiondict["sheetname"]:
        sheet_name = optiondict["sheetname"]
        excel_config.sheetname = optiondict["sheetname"]
    elif "sheetrow" in optiondict:
        sheet_name = sheet_names[optiondict["sheetrow"]]
        excel_config.sheetname = sheet_names[optiondict["sheetrow"]]
    else:
        sheet_name = sheet_names[0]
        excel_config.sheetname = sheet_names[0]

    # debugging
    if debug:
        print("sheet_name:", sheet_name)
    logger.debug("sheet_name:%s", sheet_name)

    # create a workbook sheet object - using the name to get to the right sheet
    if xlsxfiletype:
        s = wb[sheet_name]
        sheettitle = s.title
        sheetmaxrow = s.max_row
        sheetmaxcol = s.max_column
        sheetminrow = 0
        sheetmincol = 0
    else:
        s = wb.sheet_by_name(sheet_name)
        sheettitle = s.name
        sheetmaxrow = s.nrows
        sheetmaxcol = s.ncols
        sheetminrow = 0
        sheetmincol = 0

    # debugging
    if debug:
        print("sheettitle:", sheettitle)
        print("sheetmaxrow:", sheetmaxrow)
        print("sheetmaxcol:", sheetmaxcol)
    logger.debug("sheettitle:%s", sheettitle)
    logger.debug("sheetmaxrow:%s", sheetmaxrow)
    logger.debug("sheetmaxcol:%s", sheetmaxcol)

    # lower the max row limit by the record count if the record count is smaller than prior limits
    p.lower_max_row_by_reccount(sheetmaxrow)

    # check and see if we need to limit max row
    if max_rows < sheetmaxrow:
        sheetmaxrow = max_rows
        if debug:
            print("sheetmaxrow-changed:", sheetmaxrow)
            logger.debug("sheetmaxrow-changed:%s", sheetmaxrow)

    # ------------------------------- HEADER START ------------------------------

    # define the header for the records being read in
    if no_header:
        # user said we are not to look for the header in this file
        # we need to subtract 1 here because we are going to increment PAST the header
        # in the next section - so if there is no header - we need to start at zero ( -1 + 1 later)
        row_header = start_row - 1
        # 2025-01-11 changed to none as there was no row header
        row_header = None

        # 2026-03-25 - this logic was moved earlier as a configuratoin test not down here in the logic
        # if no col_aref - then we must force this to aref_result
        if False and not col_aref:
            aref_result = True
            if debug:
                print("no_header:no col_aref:set aref_result to true")
            logger.debug("no_header:no col_aref:set aref_result to true")

        # debug
        if debug:
            print("no_header:start_row:", start_row)
        logger.debug("no_header:start_row:%d", start_row)

    else:
        # fail first if we have no data
        if sheetmaxrow == 0:
            # no recordds were found - we failed
            if not allow_empty:
                # debug
                if debug:
                    print(
                        "exception:find_header:sheetmaxrow==0:no header to find"
                    )
                logger.debug(
                    "exception:find_header:sheetmaxrow==0:no header to find"
                )

                raise Exception("sheetmaxrow==0:no header to find")
            else:
                # debug
                if debug:
                    print(
                        "find_header:sheetmaxrow==0:allow_empty enabled - continue"
                    )
        # debug
        if debug:
            print("find_header:start_row:", start_row)
        logger.debug("find_header:start_row:%d", start_row)

        # look for the header in the file
        for row in range(start_row, sheetmaxrow):
            # read in a row of data
            rowdata, c_row, c_col1 = _extract_excel_row_into_list(
                xlsxfiletype, s, row, sheetmincol, sheetmaxcol, debug
            )

            # user may have specified that the first row read is the header
            if col_header:
                # first row read is header - set the values
                header = rowdata
                row_header = row
                # debugging
                if debug:
                    print("header_1strow:", header)
                logger.debug("header_1strow:%s", header)
                # validate we got a values
                header_value = [x for x in header if x]
                # if we got nothing - error out
                if not allow_empty and not header_value:
                    raise Exception(
                        "no header values found in row: "
                        + str(row)
                        + "|sheet:"
                        + str(sheet_name)
                        + "|File: "
                        + xlsfile
                    )
                # break out of this loop we are done
                break

            # have not found the header yet - so look
            if debug:
                print("looking for header at row:", row)
            logger.debug("looking for header at row:%d", row)

            # Search to see if this row is the header
            if p.matchRowList(rowdata, debug=debug) or p.search_exceeded:
                # determine if we found the header
                if p.search_exceeded:
                    # debugging
                    if debug:
                        print("exception:maxrows_search_exceeded:", p.error_msg)
                    logger.debug("maxrows in search exceeded:%s", p.error_msg)
                    # did not find the header
                    raise Exception(p.error_msg)
                elif p.search_failed:
                    # debugging
                    if debug:
                        print("exception:search_failed:", p.error_msg)
                    logger.debug("search_failed:%s", p.error_msg)
                    # did not find the header
                    raise Exception(p.error_msg)
                else:
                    # set the row_header
                    row_header = row
                    # found the header grab the output
                    header = p._data_mapped
                    # debugging
                    if debug:
                        print("header_found:", header)
                    logger.debug("header_found:%s", header)
                    # break out of the loop
                    break
            elif debug:
                print("no match found loop again")
                print("search exceeded: ", p.search_exceeded)

    # ------------------------------- HEADER END ------------------------------

    # debug
    if debug:
        print("exited header find loop")
    logger.debug("exited header find loop")

    # user wants to define/override the column headers rather than read them in
    if col_aref:
        # debugging
        if debug:
            print("copying col_aref into header")
        logger.debug("copying col_aref into header")
        # copy over the values - and determine if we need to fill in more header values
        header = col_aref[:]
        # user defined the row definiton - make sure they passed in enough values to fill the row
        if len(col_aref) < sheetmaxcol - sheetmincol:
            # not enough entries - so we add more to the end
            for _ in range(1, sheetmaxcol - sheetmincol - len(col_aref) + 1):
                header.append("")

        # now pass the final information through remapped
        header = p.remappedRow(header)
        # debug
        if debug:
            print("col_aref:header:", header)
        logger.debug("col_aref:header:%s", header)

    # ------------------------------- OBJECT DEFINITION ------------------------------
    excel_dict = {
        "xlsfile": xlsfile,
        "xlsxfiletype": xlsxfiletype,
        "keep_vba": keep_vba,
        "wb": wb,
        "sheet_names": sheet_names,
        "sheet_name": sheet_name,
        "s": s,
        "sheettitle": sheettitle,
        "sheetmaxrow": sheetmaxrow,
        "sheetmaxcol": sheetmaxcol,
        "sheetminrow": sheetminrow,
        "sheetmincol": sheetmincol,
        "row_header": row_header,
        "header": header,
        "start_row": start_row,
    }

    if debug:
        print("excel_dict: ", excel_dict)

    return excel_dict


# different name for the same function - i think this name is more meaningful
def readxls2excel_dict_findheader(
    xlsfile: str,
    req_cols: list[str],
    xlatdict: dict | None = None,
    optiondict: dict | None = None,
    col_aref: list[str] | None = None,
    data_only: bool = True,
    debug: bool = False,
) -> dict:
    return readxls_findheader(
        xlsfile,
        req_cols,
        xlatdict=xlatdict,
        optiondict=optiondict,
        col_aref=col_aref,
        data_only=data_only,
        debug=debug,
    )


# or passed on to other routines to extract the data for processing
#
# Open to edit and save:
# # example how to use:  open file for editting
# xls = kvxls.readxls_findheader( 'Wine Collection 20-05-07-v02.xlsm', [],
# optiondict={'col_header' : True}, data_only=False )
#
# # change a cell
# kvxls.setExcelCellValue( xls, 2, 'Rating', 'Changed')
# # save the file
# kvxls.writexls( xls, 'newfile.xlsm' )
#


#
# generic routine that reads in the XLS and returns back a dictionary for that xls
# that is either used to interact with that XLS object, or is passed to other routines
# that then create the dictionary/list of that xls and then close out that XLS.
#    data_only - when set to FALSE - will allow you to read macro enable file and update directly
#                and save the updated file
def chgsheet_findheader(
    excel_dict: dict,
    req_cols: list[str] | None,
    xlatdict: dict | None = None,
    optiondict: dict | None = None,
    col_aref: list[str] | None = None,
    data_only: bool = True,
    debug: bool = False,
) -> dict:
    """
    Change the current/active sheet in the XLSX that excelDict is pointed at
    based on the 'sheetname' in the 'optiondict' passed in
    """

    # test inputs
    if xlatdict is None:
        xlatdict = {}
    if optiondict is None:
        optiondict = {}
    if req_cols is None:
        req_cols = []

    # type check
    if col_aref is not None and not isinstance(col_aref, list):
        raise TypeError(f"col_aref must be type list but is: {type(col_aref)}")
    if not isinstance(req_cols, list):
        raise TypeError(f"req_cols must be type list but is: {type(req_cols)}")
    if not isinstance(optiondict, dict):
        raise TypeError(
            f"optiondict must be type dict but is: {type(optiondict)}"
        )
    if not isinstance(xlatdict, dict):
        raise TypeError(f"xlatdict must be type dict but is: {type(xlatdict)}")

    # local variables
    header = None

    # debugging
    if debug:
        print("req_cols:", req_cols)
        print("xlatdict:", xlatdict)
        print("optiondict:", optiondict)
        print("col_aref:", col_aref)
    logger.debug("req_cols:%s", req_cols)
    logger.debug("xlatdict:%s", xlatdict)
    logger.debug("optiondict:%s", optiondict)
    logger.debug("col_aref:%s", col_aref)

    # set flags
    col_header = (
        False  # if true - we take the first row of the file as the header
    )
    no_header = False  # if true - there are no headers read - we either return
    aref_result = False  # if true - we don't return dicts, we return a list
    save_row = False  # if true - then we append/save the XLSRow with the record
    save_row_abs = (
        False  # if true - then we append/save the XLSRow with the record
    )
    save_col_abs = False  # if true - then we append/save the absolute xlsx column number of the first column - from openpyxl
    save_colmap = False  # if true - then we add a new field that housed the colmapp for this
    save_colfmt = None  # if populated with a column header - than we capture and save the format of that column
    keep_vba = True  # if true - then load the xlsx with vba scripts on and save as xlsm

    start_row = 0  # if passed in - we start the search at this row (starts at 1 or greater)

    max_rows = 100000000

    # create the excel configuratoin
    excel_config = create_excel_config(
        optiondict=optiondict, func_name="kvxls.chgsheet_findheader"
    )

    # pull in passed values from optiondict
    if "col_header" in optiondict:
        col_header = optiondict["col_header"]
    if "aref_result" in optiondict:
        aref_result = optiondict["aref_result"]
    if "no_header" in optiondict:
        no_header = optiondict["no_header"]
    if "start_row" in optiondict:
        start_row = (
            optiondict["start_row"] - 1
        )  # because we are not ZERO based in the users mind
    if "save_row" in optiondict:
        save_row = optiondict["save_row"]
    if "save_row_abs" in optiondict:
        save_row_abs = optiondict["save_row_abs"]
    if "save_col_abs" in optiondict:
        save_col_abs = optiondict["save_col_abs"]
    if "save_colmap" in optiondict:
        save_colmap = optiondict["save_colmap"]
    if "save_colfmt" in optiondict:
        save_colfmt = optiondict["save_colfmt"]
    if "max_rows" in optiondict:
        max_rows = optiondict["max_rows"]
    if "keep_vba" in optiondict:
        keep_vba = optiondict["keep_vba"]

    # debugging
    if debug:
        print("chgsheet_findheader")
        print("req_cols:", req_cols)
        print("col_aref:", col_aref)
        print("col_header:", col_header)
        print("aref_result:", aref_result)
        print("no_header:", no_header)
        print("start_row:", start_row)
        print("save_row:", save_row)
        print("save_row_abs:", save_row_abs)
        print("save_col_abs:", save_col_abs)
        print("save_colmap:", save_colmap)
        print("save_colfmt:", save_colfmt)
        print("optiondict:", optiondict)
        print("excel_config:", excel_config)
    logger.debug("req_cols:%s", req_cols)
    logger.debug("col_aref%s", col_aref)
    logger.debug("col_header:%s", col_header)
    logger.debug("aref_result:%s", aref_result)
    logger.debug("no_header:%s", no_header)
    logger.debug("start_row:%s", start_row)
    logger.debug("save_row:%s", save_row)
    logger.debug("save_row_abs:%s", save_row_abs)
    logger.debug("save_row_abs:%s", save_row_abs)
    logger.debug("save_col_abs:%s", save_col_abs)
    logger.debug("save_colmap:%s", save_colmap)
    logger.debug("optiondict:%s", optiondict)

    # check to see if we are actually changing anyting - if not return back what was sent in
    if (
        "sheetname" in optiondict
        and excel_dict["sheet_name"] == optiondict["sheetname"]
    ):
        logger.debug("nothing changed - return what was sent in")
        return excel_dict

    # special condidtiaon for no header
    if not col_header and not req_cols and start_row and col_aref:
        no_header = True
        if debug:
            print(
                "Setting no_header because of col_header, start_row, col_aref"
            )
            print("no_header:", no_header)
    elif debug:
        print(f"{col_header=}\n{start_row=}\n{col_aref=}")

    # build object that will be used for record matching
    p = kvmatch.MatchRow(
        req_cols,
        xlatdict,
        optiondict,
        optiondict2={"noshowwarning": True, "fix_missing": True},
    )

    # read in values from excel_dict
    # determine what filetype we have here
    xlsfile = excel_dict["xlsfile"]
    xlsxfiletype = excel_dict["xlsxfiletype"]
    wb = excel_dict["wb"]
    sheet_names = excel_dict["sheet_names"]

    # debugging
    if debug:
        print("sheet_names:", sheet_names)
    logger.debug("sheet_names:%s", sheet_names)

    # get the sheet we are going to work with
    if "sheetname" in optiondict:
        sheet_name = optiondict["sheetname"]
    elif "sheetrow" in optiondict:
        sheet_name = sheet_names[optiondict["sheetrow"]]
    else:
        sheet_name = sheet_names[0]

    # debugging
    if debug:
        print("sheet_name:", sheet_name)
    logger.debug("sheet_name:%s", sheet_name)

    # create a workbook sheet object - using the name to get to the right sheet
    if xlsxfiletype:
        s = wb[sheet_name]
        sheettitle = s.title
        sheetmaxrow = s.max_row
        sheetmaxcol = s.max_column
        sheetminrow = 0
        sheetmincol = 0
    else:
        s = wb.sheet_by_name(sheet_name)
        sheettitle = s.name
        sheetmaxrow = s.nrows
        sheetmaxcol = s.ncols
        sheetminrow = 0
        sheetmincol = 0

    # debugging
    if debug:
        print("sheettitle:", sheettitle)
        print("sheetmaxrow:", sheetmaxrow)
        print("sheetmaxcol:", sheetmaxcol)
    logger.debug("sheettitle:%s", sheettitle)
    logger.debug("sheetmaxrow:%s", sheetmaxrow)
    logger.debug("sheetmaxcol:%s", sheetmaxcol)

    # lower the limit
    p.lower_max_row_by_reccount(sheetmaxrow)

    # check and see if we need to limit max row
    if max_rows < sheetmaxrow:
        sheetmaxrow = max_rows
        if debug:
            print("sheetmaxrow-changed:", sheetmaxrow)
            logger.debug("sheetmaxrow-changed:%s", sheetmaxrow)

    # ------------------------------- HEADER START ------------------------------

    # define the header for the records being read in
    if no_header:
        # user said we are not to look for the header in this file
        # we need to subtract 1 here because we are going to increment PAST the header
        # in the next section - so if there is no header - we need to start at zero ( -1 + 1 later)
        row_header = start_row - 1
        # 2025-01-11 changed to none as there was no row header
        row_header = None

        # if no col_aref - then we must force this to aref_result
        if not col_aref:
            aref_result = True
            if debug:
                print("no_header:no col_aref:set aref_result to true")
            logger.debug("no_header:no col_aref:set aref_result to true")

        # debug
        if debug:
            print("no_header:start_row:", start_row)
        logger.debug("no_header:start_row:%d", start_row)

    else:
        # debug
        if debug:
            print("find_header:start_row:", start_row)
        logger.debug("find_header:start_row:%d", start_row)

        # look for the header in the file
        for row in range(start_row, sheetmaxrow):
            # read in a row of data
            rowdata, c_row, c_col1 = _extract_excel_row_into_list(
                xlsxfiletype, s, row, sheetmincol, sheetmaxcol, debug
            )

            # user may have specified that the first row read is the header
            if col_header:
                # first row read is header - set the values
                header = rowdata
                row_header = row
                # debugging
                if debug:
                    print("header_1strow:", header)
                logger.debug("header_1strow:%s", header)
                # break out of this loop we are done
                break

            # have not found the header yet - so look
            if debug:
                print("looking for header at row:", row)
            logger.debug("looking for header at row:%d", row)

            # Search to see if this row is the header
            if p.matchRowList(rowdata, debug=debug) or p.search_exceeded:
                # determine if we found the header
                if p.search_exceeded:
                    # debugging
                    if debug:
                        print("maxrows_search_exceeded:", p.error_msg)
                    logger.debug("maxrows in search exceeded:%s", p.error_msg)
                    # did not find the header
                    raise Exception(p.error_msg)
                elif p.search_failed:
                    # debugging
                    if debug:
                        print("search_failed:", p.error_msg)
                    logger.debug("search_failed:%s", p.error_msg)
                    # did not find the header
                    raise Exception(p.error_msg)
                else:
                    # set the row_header
                    row_header = row
                    # found the header grab the output
                    header = p._data_mapped
                    # debugging
                    if debug:
                        print("header_found:", header)
                    logger.debug("header_found:%s", header)
                    # break out of the loop
                    break
            elif debug:
                print("no match found loop again")

    # ------------------------------- HEADER END ------------------------------

    # debug
    if debug:
        print("chgsheet_findheader:exitted header find loop")
    logger.debug("exitted header find loop")

    # user wants to define/override the column headers rather than read them in
    if col_aref:
        # debugging
        if debug:
            print("copying col_aref into header")
        logger.debug("copying col_aref into header")
        # copy over the values - and determine if we need to fill in more header values
        header = col_aref[:]
        # user defined the row definiton - make sure they passed in enough values to fill the row
        if len(col_aref) < sheetmaxcol - sheetmincol:
            # not enough entries - so we add more to the end
            for colcnt in range(
                1, sheetmaxcol - sheetmincol - len(col_aref) + 1
            ):
                header.append("")

        # now pass the final information through remapped
        header = p.remappedRow(header)
        # debug
        if debug:
            print("col_aref:header:", header)
        logger.debug("col_aref:header:%s", header)

    # ------------------------------- OBJECT DEFINITION ------------------------------
    excel_dict = {
        "xlsfile": xlsfile,
        "xlsxfiletype": xlsxfiletype,
        "keep_vba": keep_vba,
        "wb": wb,
        "sheet_names": sheet_names,
        "sheet_name": sheet_name,
        "s": s,
        "sheettitle": sheettitle,
        "sheetmaxrow": sheetmaxrow,
        "sheetmaxcol": sheetmaxcol,
        "sheetminrow": sheetminrow,
        "sheetmincol": sheetmincol,
        "row_header": row_header,
        "header": header,
        "start_row": start_row,
    }

    if debug:
        print("excel_dict: ", excel_dict)

    return excel_dict


# ---------- EXTRACT DATA FROM EXCEL  ----------------------
#
# coding structure - build one generic (INTERNAL) function that does all the various things
# with passed in variables that are all optional
# and based on variable settings - executes the behavior being asked
#
# then create external functions - with clear passed in parameters
# that calls this internal function with teh right settings
#

# read in the CSV and create a dictionary to the records
# based on one or more key fields
# assumes the first line of the CSV file is the header/defintion of the CSV
#
# break_blank_row - when you encounter a blank row - stop reading rows
# skip_blank_row - skip rows that are blank but process all rows
#
#
# features to add
#
#   noheader - flag means we pass in the array that is the header
#   col_header - flag means we get the header from the very first line (or start_row) in the file
#   aref_result - flag that tells us to return the array of array (otherwise we pass back the array of dict)
#
#   col_aref - array that defines the columns we read into dict
#
#   start_row - user entered values start at row 1 (not zero)
#
#   unique_column - if enabled, we must validate that our column names are unique
#   ignore_blank_row - if enabled, if the row read has no data - we don't put the data into the extracted list
#   ignore_not_fill - not sure what this one is
#
#   dateflds - array of fields in xls that convert to a date
#   save_row - flag defines if we should save the row number
#   save_row_abs - flag defines if we should save the row number
#   save_col_abs - flag defines if we should save the row number
#   save_colmap - flag defines if we should create the column mapping column
#
#   required_fields_populated - checking logic to assure that all required fields have data with optional
#     required_fld_swap - a dict that says if key is not populated - check the value
#                         tied to that key to see if it is populated
#


def readxls2list_findheader(
    xlsfile: str | os.PathLike,
    req_cols: list[str],
    xlatdict: dict | None = None,
    optiondict: dict | None = None,
    col_aref: list[str] | None = None,
    debug: bool = False,
) -> list[dict]:
    if xlatdict is None:
        xlatdict = {}
    if optiondict is None:
        optiondict = {}
    if req_cols is None:
        req_cols = []

    # local variables
    # results = []
    # header = None

    # debugging
    if debug:
        print("req_cols:", req_cols)
        print("xlatdict:", xlatdict)
        print("optiondict:", optiondict)
        print("col_aref:", col_aref)
    logger.debug("req_cols:%s", req_cols)
    logger.debug("xlatdict:%s", xlatdict)
    logger.debug("optiondict:%s", optiondict)
    logger.debug("col_aref:%s", col_aref)

    # set flags
    # col_header = False  # if true - we take the first row of the file as the header
    # no_header = False  # if true - there are no headers read - we either return
    # aref_result = False  # if true - we don't return dicts, we return a list
    # save_row = False  # if true - then we append/save the XLSRow with the record
    # save_row_abs = False  # if true - then we append/save the XLSRow with the record
    # save_col_abs = False  # if true - then we append/save the XLSRow with the record
    # save_colmap = False # if true - then we add a new field that housed the colmapp for this
    # save_colfmt = None # if populated with a column header - than we capture and save the format of that column

    # start_row = 0  # if passed in - we start the search at this row (starts at 1 or greater)

    # call the routine that opens the XLS and returns back the excel_dict
    # (missing data_only attribute between optiondict and debug)
    excel_dict = readxls_findheader(
        xlsfile, req_cols, xlatdict, optiondict, col_aref, debug=debug
    )

    # call the library function
    return excelDict2list_findheader(
        excel_dict,
        req_cols,
        xlatdict=xlatdict,
        optiondict=optiondict,
        col_aref=col_aref,
        debug=debug,
    )


def excelDict2list_findheader(
    excel_dict: dict,
    req_cols: list[str] | None = None,
    xlatdict: dict | None = None,
    optiondict: dict | None = None,
    col_aref: list[str] | None = None,
    debug: bool = False,
) -> list[dict]:

    if optiondict is None:
        optiondict = dict()
    if xlatdict is None:
        xlatdict = dict()
    if col_aref is None:
        col_aref = list()

    # test inputs
    if req_cols and not isinstance(req_cols, list):
        raise TypeError(f"req_cols must be type list but is: {type(req_cols)}")
    if not isinstance(col_aref, list):
        raise TypeError(f"col_aref must be type list but is: {type(col_aref)}")

    # local variables
    results = []
    header = None

    # debugging
    if debug:
        print("req_cols:", req_cols)
        print("xlatdict:", xlatdict)
        print("optiondict:", optiondict)
        print("col_aref:", col_aref)
    logger.debug("req_cols:%s", req_cols)
    logger.debug("xlatdict:%s", xlatdict)
    logger.debug("optiondict:%s", optiondict)
    logger.debug("col_aref:%s", col_aref)

    # set flags
    col_header = (
        False  # if true - we take the first row of the file as the header
    )
    no_header = False  # if true - there are no headers read - we either return
    aref_result = False  # if true - we don't return dicts, we return a list
    save_row = False  # if true - then we append/save the XLSRow with the record
    save_row_abs = (
        False  # if true - then we append/save the XLSRow with the record
    )
    save_col_abs = False  # if true - then we append/save the absolute xlsx column number of the first column - from openpyxl
    save_colmap = False  # if true - then we add a new field that housed the colmapp for this
    save_colfmt = None  # if populated with a column header - than we capture and save the format of that column

    start_row = 0  # if passed in - we start the search at this row (starts at 1 or greater)

    # pull in passed values from optiondict
    if "col_header" in optiondict:
        col_header = optiondict["col_header"]
    if "aref_result" in optiondict:
        aref_result = optiondict["aref_result"]
    if "no_header" in optiondict:
        no_header = optiondict["no_header"]
    if "start_row" in optiondict:
        start_row = (
            optiondict["start_row"] - 1
        )  # because we are not ZERO based in the users mind
    if "save_row" in optiondict:
        save_row = optiondict["save_row"]
    if "save_row_abs" in optiondict:
        save_row_abs = optiondict["save_row_abs"]
    if "save_col_abs" in optiondict:
        save_col_abs = optiondict["save_col_abs"]
    if "save_colmap" in optiondict:
        save_colmap = optiondict["save_colmap"]
    if "save_colfmt" in optiondict:
        save_colfmt = optiondict["save_colfmt"]

    # debugging
    if debug:
        print("col_header:", col_header)
        print("aref_result:", aref_result)
        print("no_header:", no_header)
        print("start_row:", start_row)
        print("save_row:", save_row)
        print("save_row_abs:", save_row_abs)
        print("save_col_abs:", save_col_abs)
        print("save_colmap:", save_colmap)
        print("save_colfmt:", save_colfmt)
        print("optiondict:", optiondict)
    logger.debug("col_header:%s", col_header)
    logger.debug("aref_result:%s", aref_result)
    logger.debug("no_header:%s", no_header)
    logger.debug("start_row:%s", start_row)
    logger.debug("save_row:%s", save_row)
    logger.debug("save_row_abs:%s", save_row_abs)
    logger.debug("save_col_abs:%s", save_col_abs)
    logger.debug("save_colmap:%s", save_colmap)
    logger.debug("save_colfmt:%s", save_colfmt)
    logger.debug("optiondict:%s", optiondict)

    # expand out all the values that came from excel_dict
    xlsxfiletype = excel_dict["xlsxfiletype"]
    # wb = excel_dict['wb']
    # sheet_names = excel_dict['sheet_names']
    # sheet_name = excel_dict['sheet_name']
    s = excel_dict["s"]
    sheettitle = excel_dict["sheettitle"]
    sheetmaxrow = excel_dict["sheetmaxrow"]
    sheetmaxcol = excel_dict["sheetmaxcol"]
    # sheetminrow = excel_dict['sheetminrow']
    sheetmincol = excel_dict["sheetmincol"]
    row_header = excel_dict["row_header"]
    header = excel_dict["header"]
    start_row = excel_dict["start_row"]

    # if we don't have a header we must set the aref_result flag
    if not header and not aref_result:
        if debug:
            print("setting aref_results because there is no header")
        logger.debug("setting aref_results becaus there is no header")

        aref_result = True

    # if we dont' have a row_header then use start_row
    if row_header is None:
        row_data_start = start_row
    else:
        row_data_start = row_header + 1

    # debugging
    if debug:
        print("sheettitle:", sheettitle)
        print("sheetmaxrow:", sheetmaxrow)
        print("sheetmaxcol:", sheetmaxcol)

    # ------------------------------- RECORDS START ------------------------------

    for row in range(row_data_start, sheetmaxrow):
        # read in a row of data
        rowdata, c_row, c_col1 = _extract_excel_row_into_list(
            xlsxfiletype, s, row, sheetmincol, sheetmaxcol, debug
        )

        # break on blank row
        if "break_blank_row" in optiondict and optiondict["break_blank_row"]:
            non_empty = [x for x in rowdata if x]
            if not non_empty:
                if debug:
                    print("break blank row:", row + 1, ":", rowdata)
                break

        # skip on blank row
        if "skip_blank_row" in optiondict and optiondict["skip_blank_row"]:
            non_empty = [x for x in rowdata if x]
            if not non_empty:
                if debug:
                    print("skip blank row:", row + 1, ":", rowdata)
                continue

        # determine what we are returning
        if aref_result:
            # we want to return the data we read
            rowdict = rowdata
            if debug:
                print("saving as array")
            logger.debug("saving as array")

            # optionally add the XLSRow attribute to this dictionary (not here right now
            if save_row:
                rowdict.append(row + 1)
                if debug:
                    print("append row to record")
                logger.debug("append row to record")

            # optionally add the XLSRow attribute to this dictionary (not here right now
            if save_row_abs:
                rowdict.append(c_row)
                if debug:
                    print("append row to record")
                logger.debug("append row to record")

            # optionally add the XLSRow attribute to this dictionary (not here right now
            if save_col_abs:
                rowdict.append(c_col1)
                if debug:
                    print("append col1 to record")
                logger.debug("append col1 to record")

                # TODO - put colmap logic here

        else:
            if debug:
                print("saving as dict")
                print("header:", header)
                print("rowdata:", rowdata)
            logger.debug("saving as dict:header:%s:rowdata:%s", header, rowdata)

            # we found the header so now build up the records
            rowdict = dict(zip(header, rowdata))

            # optionally add the XLSRow attribute to this dictionary (not here right now
            if save_row:
                rowdict[FLD_XLSROW] = row + 1
                if debug:
                    print("add column XLSRow with row to record")
                logger.debug("add column XLSRow with row to record")

            # optionally add the XLSRow attribute to this dictionary (not here right now
            if save_row_abs:
                rowdict[FLD_XLSROW_ABS] = c_row
                if debug:
                    print("add column XLSRowAbs with row to record")
                logger.debug("add column XLSRowAbs with row to record")

            # optionally add the XLSRow attribute to this dictionary (not here right now
            if save_col_abs:
                rowdict[FLD_XLSCOL_ABS] = c_col1
                if debug:
                    print("add column XLSCol1 with row to record")
                logger.debug("add column XLSCol1 with row to record")

            # TODO put the colmap logic here

            # put in the column formatting here
            if save_colfmt:
                # grab the pattern
                cell_color, cell_fill_type, cell_start_color, cell_end_color, cell_fill = (
                    getExcelCellPatternFill(
                        excel_dict, row, save_colfmt, debug=debug
                    )
                )
                # grab the font
                (
                    cell_font_name,
                    cell_font_size,
                    cell_font_bold,
                    cell_font_italic,
                    cell_font_underline,
                    cell_font_strike,
                    cell_font_color,
                ) = getExcelCellFont(excel_dict, row, save_colfmt, debug=debug)

                # make a dict that we can convert to json
                colfmt_dict = {
                    "cell_color": cell_color,
                    "cell_fill_type": cell_fill_type,
                    "cell_start_color": cell_start_color,
                    "cell_end_color": cell_end_color,
                    "cell_font_name": cell_font_name,
                    "cell_font_size": cell_font_size,
                    "cell_font_bold": cell_font_bold,
                    "cell_font_italic": cell_font_italic,
                    "cell_font_underline": cell_font_underline,
                    "cell_font_strike": cell_font_strike,
                    "cell_font_color": cell_font_color,
                }
                # save this
                rowdict[FLD_XLSFMT] = colfmt_dict

            # do field manipulations here - date - but only on XLS not XLSX files
            if not xlsxfiletype:
                if "dateflds" in optiondict:
                    for fld in optiondict["dateflds"]:
                        if fld in rowdict:
                            rowdict[fld] = xldate_to_datetime(rowdict[fld])
                            if debug:
                                print("xldate conversion on:", fld)
                            logger.debug("xldate conversion on:%s", fld)

        # add this dictionary to the results
        results.append(rowdict)
        if debug:
            print("append rowdict to results")
        logger.debug("append rowdict to results")

    # ------------------------------- RECORDS END ------------------------------

    # debugging
    # if debug: print('results:', results)

    # return the results
    return results


# read in the XLS and create a dictionary to the records
# based on one or more key fields
def readxls2dict_findheader(
    xlsfile: str,
    dictkeys: list[str],
    req_cols: list[str] | None = None,
    xlatdict: dict | None = None,
    optiondict: dict | None = None,
    col_aref: list[str] | None = None,
    dupkeyfail: bool = False,
    debug: bool = False,
) -> dict:
    if xlatdict is None:
        xlatdict = {}
    if optiondict is None:
        optiondict = {}
    if req_cols is None:
        req_cols = []

    # validate we have proper input
    if not dictkeys:
        raise ValueError("dictkeys must be populated")

    # test how dictkeys was passed in
    if isinstance(dictkeys, str):
        dictkeys = [dictkeys]
        if debug:
            print(
                "readxls2dict_findheader:converted dictkeys from string to list"
            )
        logger.debug("converted dictkeys from string to list")

    # validate inputs
    if not isinstance(dictkeys, list):
        raise TypeError("dictkeys must be list but is: {type(dictkeys)}")
    if not isinstance(req_cols, list):
        raise TypeError("req_cols must be list but is: {type(req_cols)}")
    if not isinstance(xlatdict, dict):
        raise TypeError("xlatdict must be dict but is: {type(xlatdict)}")
    if not isinstance(optiondict, dict):
        raise TypeError("optiondict must be dict but is: {type(optiondict)}")

    # check for duplicate keys
    dupkeys = []

    # results defined as a dicut
    results = {}

    # debugging
    logger.debug("dictkeys:%s", dictkeys)
    if debug:
        print("readxls2dict_findheader:dictkeys:", dictkeys)
        print("readxls2dict_findheader:reading in xls as a list first")
    logger.debug("reading in xls as a list first")

    # read in the data from the file
    resultslist = readxls2list_findheader(
        xlsfile,
        req_cols,
        xlatdict=xlatdict,
        optiondict=optiondict,
        col_aref=col_aref,
        debug=debug,
    )

    # debugging
    if debug:
        print(
            "readxls2dict_findheader:xls data is in an array - now convert to a dictionary"
        )
        print("readxls2dict_findheader:dictkeys:", dictkeys)
    logger.debug("xls data is in an array - now convert to a dictionary")
    logger.debug("dictkeys:%s", dictkeys)

    # convert to a dictionary based on keys provided
    for rowdict in resultslist:
        # rowdict = dict(zip(header,row))
        if debug:
            print("rowdict:", rowdict)
            print("dictkeys:", dictkeys)
        logger.debug("rowdict:%s", rowdict)
        logger.debug("dictkeys:%s", dictkeys)
        reckey = kvmatch.build_multifield_key(rowdict, dictkeys)
        # do we fail if we see the same key multiple times?
        if dupkeyfail:
            if reckey in results.keys():
                # capture this key
                dupkeys.append(reckey)

        # create/update the dictionary
        results[reckey] = rowdict

    # fail if we found dupkeys
    if dupkeys:
        logger.error("duplicate key failure:%s", ",".join(dupkeys))
        print("readxls2dict:duplicate key failure:", ",".join(dupkeys))
        raise

    # return the results
    return results


# -------- WRITE FILES -------------------------


# write out a dict of (dict or aref) to an XLS/XLSX based on the filename passed in
def writedict2xls(
    xlsfile: str,
    data: dict,
    col_aref: list[str] | None = None,
    optiondict: dict | None = None,
    debug: bool = False,
):
    # convert dict to array and then call writelist2xls
    if not data:
        data2 = None
    else:
        data2 = [data[key] for key in sorted(data.keys())]

    # call the other library
    return writelist2xls(
        xlsfile, data2, col_aref=None, optiondict=None, debug=debug
    )


# write out a list of (dict or aref) to an XLS/XLSX based on the filename passed in
def writelist2xls(
    xlsfile: str,
    data: list[dict],
    col_aref: list[str] | None = None,
    optiondict: dict | None = None,
    debug: bool = False,
):
    """
    Create or update XLS/XLSX with a list of values that are written to a sheet

    Inputs:
        xlsfile - str - filename/path to the file being created/updated
        data - list of list or list of dict - this is the data being written out
        col_aref - list - column names to be used as the header
        optiondict - dict - list of options to be processed (defined below)
        debug - bool - when true - display processing messages

    Returns:


    optiondict:
        sheet_name - defines the sheet_name you are creating in this xlsx
        replace_sheet - we are adding/inserting a sheet into an exising file if one exists or creating the file
        replace_index - if we want to position the new sheet- we can defiune where we want it
                    (0 is first sheet, -1 is last sheet, no value is last sheet)
        start_row - the row we start the output on


    """
    if optiondict is None:
        optiondict = dict()
    if col_aref is None:
        col_aref = list()

    # test inputs
    if not isinstance(optiondict, dict):
        raise TypeError(f"optiondict must be dict but is: {type(optiondict)}")
    if not isinstance(col_aref, list):
        raise TypeError(f"col_aref must be list but is: {type(col_aref)}")

    # debugging
    if debug:
        print("writelist2xls")
        print("xlsfile:", xlsfile)
        print("col_aref:", col_aref)
        print("optiondict:", optiondict)

    # local variables default values
    sheetname = "Sheet1"
    no_header = False
    aref_result = False
    replace_sheet = False
    replace_index = None

    # create the excel configuratoin
    excel_config = create_excel_config(
        optiondict=optiondict, func_name="kvxls.writelist2xls"
    )

    # check and override configuration
    if not excel_config.sheetname:
        excel_config.sheetname = sheetname

    # determine what filetype we have here
    xlsxfiletype = xlsfile.endswith(".xlsx") or xlsfile.endswith(".xlsm")
    excel_config.xlsxfiletype = xlsxfiletype

    # change settings based on user input
    if "sheetname" in optiondict:
        sheetname = optiondict["sheetname"]
    if "no_header" in optiondict:
        no_header = optiondict["no_header"]
    if "aref_result" in optiondict:
        aref_result = optiondict["aref_result"]
    if "replace_sheet" in optiondict:
        replace_sheet = optiondict["replace_sheet"]
    if "replace_index" in optiondict:
        replace_index = optiondict["replace_index"]

    # no data passed in - set up to create an empty file
    if not data:
        aref_result = True
        excel_config.aref_result = True
        if not isinstance(data, list):
            data = list()
    else:
        # if we set aref_result and the record we pass in is dict, overwrite the flag
        if aref_result and isinstance(data[0], dict):
            aref_result = False
            excel_config.aref_result = False
        # set this value if the record we get is a list not a dictionary
        if isinstance(data[0], list):
            aref_result = True
            excel_config.aref_result = True

    # debugging
    if debug:
        print("sheetname:", sheetname)
        print("no_header:", no_header)
        print("aref_result:", aref_result)
        print("replace_sheet:", replace_sheet)
        print("replace_index:", replace_index)
        print("xlsxfiletype:", xlsxfiletype)
        print("data cnt:", len(data))
        print("excel_config: ", excel_config)

    # validate we have columns defined - or create one if we can
    if not col_aref:
        if aref_result:
            # this is a list passed in - we don't need header
            no_header = True
            excel_config.no_header = True
        else:
            # we can pull the keys from this record to create the col_aref
            col_aref = list(data[0].keys())
            excel_config.col_aref = list(data[0].keys())

    # validate we have the right type of variable
    if col_aref and not isinstance(col_aref, list):
        raise TypeError(f"col_aref must be list but is: {type(col_aref)}")

    # debuging
    if debug:
        print("col_aref:", col_aref)

    # Create a new workbook
    if xlsxfiletype:
        # XLSX file
        if replace_sheet and sheetname and os.path.exists(xlsfile):
            # we want to replace/add a sheet to an existing file
            # we set replace_sheet True, we specified a sheetname
            # and the file exists
            if debug:
                print("read in the file with openpyxl to add/replace a sheet")

            # we are performing a replace/insert of a sheet in an existing workbook
            # open the workbook
            wb = openpyxl.load_workbook(xlsfile)
            # get the list of sheets that already exist
            sheets = wb.sheetnames
            if sheetname in sheets:
                # delete the sheet if it already exists (this is a replace)
                del wb[sheetname]
            if replace_index is None:
                # did not specifiy where we want this sheet - so just create the heet
                ws = wb.create_sheet(sheetname)
            else:
                # specified where we want this sheet to create it in the positoin of interest
                ws = wb.create_sheet(sheetname, replace_index)
        else:
            # file did not exist, or we did not set replace_sheet True
            if debug:
                print("creating new workbook")

            # open a new object
            wb = openpyxl.Workbook()
            # get the new sheet
            ws = wb.active

        # set the title if one is specified
        if sheetname != "Sheet1":
            ws.title = sheetname

    elif XLSXONLY:
        # put in to deal with a simplified library
        raise NotImplementedError(
            "this library only supports newer Excel file types"
        )
    else:
        # XLS file - create the output work book we want to create
        if replace_sheet and sheetname and os.path.exists(xlsfile):
            if debug:
                print("read in the file with xlrd")

            # we are performing a replace/insert of a sheet in an existing workbook
            # read in the origianl file
            wbin = xlrd.open_workbook(xlsfile, formatting_info=True)

            # get list of sheets
            sheetsin = wbin.sheet_names()
            # debugging
            if debug:
                print("xlsfile:", xlsfile)
                print("sheetsin:", sheetsin)
                if sheetname in sheetsin:
                    print("need to remove:", sheetname)

            # copy over
            wb = xl_copy(wbin)
            if debug:
                print("Copy read in data to write out work book")

            # special processing if the new sheetname already exists
            if sheetname in sheetsin:
                # get the list of sheets in this output
                wb_sheets = wb._Workbook__worksheets

                # remove sheet if it exists already
                for sheet in wb_sheets:
                    # capture the sheet we need to remove
                    if sheetname == sheet.name:
                        wb_sheets.remove(sheet)
                        if debug:
                            print("xwlt sheet removed:", sheetname)

                # take this final list
                wb._Workbook__worksheets = wb_sheets
                if debug:
                    print("copied the remaining wb_sheets to replace wb")
                    for sheet in wb._Workbook__worksheets:
                        print("sheet.name:", sheet.name)

                # save this strippped file
                wb.save(xlsfile)
                if debug:
                    print("saved out file:", xlsfile)

                # read in and copy
                wbin = xlrd.open_workbook(xlsfile, formatting_info=True)
                wb = xl_copy(wbin)
                wb_sheets = wb._Workbook__worksheets

                if debug:
                    print("Sheets from saved and reloaded file")
                    for sheet in wb._Workbook__worksheets:
                        print("sheet.name:", sheet.name)
            elif debug:
                print(
                    "Sheet does not exist - so no special processing takes place:",
                    sheetname,
                )

        else:
            if debug:
                print("new work book with xlwt")
            wb = xlwt.Workbook()  # None # xlrd.open_workbook(xlsfile)

        # now add the sheet
        ws = wb.add_sheet(sheetname, cell_overwrite_ok=True)

    # set the output row
    xlsrow = 0
    if "start_row" in optiondict and optiondict["start_row"]:
        xlsrow = optiondict["start_row"] - 1

    # get the header created
    if not no_header:
        # put column text in each column of this header row
        for xlscol in range(0, len(col_aref)):
            if xlsxfiletype:
                ws.cell(
                    row=xlsrow + 1, column=xlscol + 1, value=col_aref[xlscol]
                )
            else:
                ws.write(xlsrow, xlscol, col_aref[xlscol])

        # increment the row
        xlsrow += 1

    # now step through the data itself
    for record in data:
        if debug:
            print(record)

        # output this row of data
        if col_aref and len(col_aref):
            for xlscol in range(0, len(col_aref)):
                # determine the value - based on how the records are structured
                try:
                    if aref_result:
                        # record is a list
                        value = record[xlscol]
                    else:
                        # record is a dict - see if we can grab this key
                        # and if not return an empty string
                        value = record.get(col_aref[xlscol], "")
                except Exception as e:
                    value = ""
                    if debug:
                        print("kvxls-set value failed with error: ", e)

                # could put a feature in here to convert the value to a string before storing
                if xlsxfiletype:
                    # display messages when we get unexpected types to be output
                    if not isinstance(
                        value,
                        (str, int, float, bool, datetime.datetime, NoneType),
                    ):
                        print(
                            "kvxls.writelist2xls - value not of standard type"
                        )
                        print(f"type: {type(value)}")
                        print(f"value: {value}")
                        print(f"xlsfile: {xlsfile}")
                        print(f"xlscol: {xlscol}")
                        if isinstance(record, dict):
                            print(f"col: {list(record.keys())[xlscol]}")
                        print(f"record: {record}")
                    ws.cell(row=xlsrow + 1, column=xlscol + 1, value=value)
                else:
                    ws.write(xlsrow, xlscol, value)
        elif aref_result:
            # no header being output - we are processing a list not a dict
            for xlscol in range(0, len(record)):
                # get value by index number
                value = record[xlscol]

                if xlsxfiletype:
                    # display messages when we get unexpected types to be output
                    if not isinstance(
                        value,
                        (str, int, float, bool, datetime.datetime, NoneType),
                    ):
                        print(
                            "kvxls.writelist2xls - value not of standard type"
                        )
                        print(f"type: {type(value)}")
                        print(f"value: {value}")
                        print(f"xlsfile: {xlsfile}")
                        print(f"xlscol: {xlscol}")
                        if isinstance(record, dict):
                            print(f"col: {list(record.keys())[xlscol]}")
                        print(f"record: {record}")
                    ws.cell(row=xlsrow + 1, column=xlscol + 1, value=value)
                else:
                    ws.write(xlsrow, xlscol, value)

        # done with this row - increment counter
        xlsrow += 1

    if debug:
        print("saving file, sheet:", xlsfile, sheetname)

    # now save this object
    wb.save(xlsfile)

    # return the filename we saved it as
    return xlsfile


# write out a XLSX object in memory
def writexls(
    excel_dict: dict,
    xlsfile: str | None = None,
    xlsm: bool = False,
    debug: bool = False,
):
    """
    Save the current excelDict object to a file - this only works for XLSX files

    Change the file extensoin if we are saving with vba

    Inputs:
        excel_dict - dict - the object housing hte XLSX defintion
        xlsfile - str - the filename to save as, if not populated we use the filename that created excel_dict

    """
    # test inputs
    if not isinstance(excel_dict, dict):
        raise TypeError("excel_dict must be dict but is: {type(excel_dict)}")

    # check to see that we can do this
    if not excel_dict["xlsxfiletype"]:
        raise NotImplementedError(
            "feature not supported for XLS files only XLSX"
        )

    # if the user did not pass in a filename
    # use the same filename we read in
    if not xlsfile:
        xlsfile = excel_dict["xlsfile"]

    # change the file extention to xlsm if flag is set
    if xlsm or excel_dict["keep_vba"]:
        if debug:
            print("Changing filename from: ", xlsfile)
        filename, file_ext = os.path.splitext(xlsfile)
        xlsfile = filename + ".xlsm"

    # debugging
    if debug:
        print("Saving to: ", xlsfile)

    # get the workbook
    wb = excel_dict["wb"]

    # now save this object
    wb.save(xlsfile)

    # return the filename just saed
    return xlsfile


if __name__ == "__main__":
    # put some quick test code here
    pass

# eof
