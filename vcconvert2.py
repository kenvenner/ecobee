"""
@author:   Ken Venner
@contact:  ken@venerllc.com
@version:  1.28

Read information from Beautiful Places XLS files,
extract out occupancy data, build a new
output file that has an entry per day of stay and stay type
"""

import pprint
pp = pprint.PrettyPrinter(indent=4)

# we are reusing features from another applicatoin but we wnat
# the log files to be tied to this applicatoin - so we call tihs
# application/library late in order to have the logger ocnfigured to THIS app
import villaecobee
import villacalendar

import kvutil
import kvxls
import kvcsv

import copy
import os
import shutil

import datetime
import sys

import pool

# for sorting a list of dicts
from operator import itemgetter

# working with Excel files
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import numbers
from openpyxl.utils import get_column_letter

# Excel formatting strings
bold = Font(bold=True, name="Arial", size=10)
regular = Font(name="Arial", size=10)
fit_centered = Alignment(shrink_to_fit=True, horizontal="center")
fit = Alignment(shrink_to_fit=True)

# this utility is used to convert the Beautiful Places Villa bookings XLS
# file into the flattened text file used by the "villaecobee.py" file
#
# General process:
# 1) get a new xls from beautiful places
# 2) save it to the directory/filename where this tool is stored
# 3) run this tool:  python vcconvert.py


import kvlogger
config=kvlogger.get_config(kvutil.filename_create(__file__, filename_ext='log', path_blank=True), loggerlevel='INFO') #single file
kvlogger.dictConfig(config)
logger=kvlogger.getLogger(__name__)

# application variables
optiondictconfig = {
    'AppVersion': {
        'value': '1.28',
        'description': 'defines the version number for the app',
    },
    'debug': {
        'value': False,
        'type': 'bool',
        'description': 'defines if we are running in debug mode',
    },
    'verbose': {
        'value': 1,
        'type': 'int',
        'description': 'defines the display level for print messages',
    },
    'xls_filename': {
        'value': 'Attune_Estate_2025_Bookings.xlsx',
        'description': 'defines the name of the BP xls filename',
    },
    'occupy_alt_dir': {
        'value': 'g:/My Drive/VillaRaspi/',
        'description': 'defines the directory where we put a copy of occupy_filename',
    },
    'occupy_filename': {
        'value': 'stays.txt',
        'description': 'defines the name of the file holding the villa occupancy',
    },
    'pool_heater_allowed_alt_dir': {
        'value': 'g:/My Drive/VillaRaspi/pool_heater_allowed/',
        'description': 'defines the directory where we put a copy of pool_heater_allowed_filename'
    },
    'pool_heater_allowed_filename': {
        'value': 'pool_heater_allowed.txt',
        'description': 'defines the name of the file holding the days the pool heater is allowed',
    },
    'occupy_history_filename': {
        'value': 'stays_history.txt',
        'description': 'defines the name of the file holding the historical villa occupancy',
    },
    'xlsdateflds': {
        'value': ['First Night', 'Checkout Day', 'BookedOn', 'HoldUntil'],
        'type': 'liststr',
        'description': 'defines the list of date fields inside the xls',
    },
    'fldFirstNight': {
        'value': 'First Night',
        'description': 'defines the name of the field holding the first night date',
    },
    'fldNights': {
        'value': 'Nights',
        'description': 'defines the name of the field holding the int of nights stay',
    },
    'fldLastNight': {
        'value': 'Checkout Day',
        'description': 'defines the name of the field holding the first night date',
    },
    'fldType': {
        'value': 'Type',
        'description': 'defines the name of the field holding the type of the stay',
    },
    'fldDate': {
        'value': 'date',
        'description': 'defines the name of the field that holds the date in the occupancy file',
    },
    'calendarsync': {
        'type': 'bool',
        'value': True,
        'description': 'defines if we are going to sync XLS data with calendar',
    },
    'startback': {
        'type': 'int',
        'description': 'defines number of days added to today that we update the calendar (negative numbers are in the past)',
    },
    'startdate': {
        'value': None,
        'type': 'date',
        'description': 'defines the start date, use this field or startback field to change the startdate',
    },
}

### GLOBAL VARIABLES ####

# date information
ADD_ONE_DAY = datetime.timedelta(days=1)
ADD_ONE_WEEK = datetime.timedelta(days=7)
ADD_TWO_WEEK = datetime.timedelta(days=14)
DATE_FMT = '%m/%d/%Y'

# xls file occtype conversion to
# an array that is:  new code and # of days to add to stay for temp control
OCC_TYPE_CONV = {
    'Hold - Clean': ['C', 0],
    'Hold - Construction': ['O', 1],
    'Hold - Fall Mbr Event': ['O', 1],
    'Hold - Harvest Party': ['O', 1],
    'Hold - Maint': ['M', 0],
    'Hold - Maint.': ['M', 0],
    'Hold - Other': ['M', 0],
    'Hold - Owner Scribner': ['O', 1],
    'Hold - Owner': ['O', 1],
    'Hold - Release Party': ['O', 1],
    'Hold - Renter': ['R', 1],
    'Hold - Scribner': ['O', 1],
    'Hold - Venner': ['O', 1],
    'Hold - Winery Business': ['O', 0],
    'Hold- Mainten': ['M', 0],
    'Hold-Deep Clean': ['M', 0],
    'Hold-Owner': ['O', 1],
    'Hold-Renter': ['R', 1],
    'Karli Vendor Tour': ['O', 1],
    'Owners Hold - Harvest': ['O', 1],
    'Owners Hold - Venner': ['O', 1],
    'Owners Hold - Scribner': ['O', 1],
    'Owners Hold': ['O', 1],
    'Owners Hold- Venner': ['O', 1],
    'Res - Owner': ['O', 1],
    'Res - Renter': ['R', 1],
    'Res- Renter': ['R', 1],
    'Res-Renter': ['R', 1],
    'Res. - Owner': ['O', 1],
    'Res. - Renter': ['R', 1],
    'Rest - Renter': ['R', 1],
    'Res.-Owner': ['O', 1],
    'Res.-Renter': ['R', 1],
    'VRBO/Repeat guest': ['R',1],

}

OCC_TYPE_2_BOOKING_CODE = {
    'O': 'OWN',
    'R': 'MLS',
    'C': 'CLN',
    'M': 'MSM',
}

# field that holds the booking ids
BOOKING_FLD = 'Booking'
REVPERDAY_FLD = "RevPerDay"
REVTOTAL_FLD = 'Rent'
STAYS_FLD = 'Nights'
POOL_FLD = 'PoolEnabled'


# xls header definition
COL_REQUIRED = [
    BOOKING_FLD,
    'First Night',
    'Checkout Day',
    STAYS_FLD,
    'Type',
    REVTOTAL_FLD,
    'Source',
    'HoldUntil']

# old fields
OLD_FIELDS = [
    'Managing',
    'Confirmed',
    'BookedOn',
]

COL_CENTERED = ['Nights', 'Source', 'Managing', 'Confirmed']

COL_NUMBER_FLDS = ['Rent']

COL_WIDTH = {
    BOOKING_FLD: 12,
    'First Night': 12,
    'Checkout Day': 15,
    'Type': 17,
    REVTOTAL_FLD: 12,
    'Source': 17,
    'Managing': 15,
    'Confirmed': 12,
    'BookedOn': 12,
    'HoldUntil': 12,
    REVPERDAY_FLD: 12,
}

MON_STRING = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

SHEET_LISTING = 'Listing'

def validate_res_records(xlsaref, fldFirstNight, fldNights, fldLastNight, fldType):
    """
    Step through each record and make sure the record is considered valid or generate an error mesage

    :param xlsaref: (list of dicts) recodrds from that file from sheet "Listing"
    :param fldFirstNight: (str) column header of the first night date column
    :param fldNights: (str) column header for the field holding the int # of nights
    :param fldLastNight: (str) column header of the last night date column
    :param fldType: (str) column header of the reservation type column

    :return errors: (list) - list of errors we found
    """
    errors = list()
    booking = dict()
    for recidx, rec in enumerate(xlsaref):
        date_error = False
        # date fields are dates
        for fld in (fldFirstNight, fldLastNight):
            if not isinstance(rec[fld], datetime.datetime):
                date_error = True
                errors.append('Field {} not of type datetime: {}'.format(fld,
                                                                         {'recidx': recidx,
                                                                          'type': type(rec[fldFirstNight]),
                                                                          'rec': rec}))

        # if no errors - make sure the difference matches the set in XLSX
        if not date_error:
            dt_diff = rec[fldLastNight] - rec[fldFirstNight]
            num_nights = int(rec[fldNights])
            if dt_diff.days != num_nights:
                errors.append('Field {} not calc as date difference: {}'.format(fld,
                                                                                {'recidx': recidx,
                                                                                 'dt_diff': dt_diff.days,
                                                                                 'num_nights': num_nights,
                                                                                 'rec': rec}))

        # check the record / reservation type
        if rec[fldType] not in OCC_TYPE_CONV:
            errors.append('Field {} not in OCC_TYPE_CONV: {}'.format(fldType,
                                                                     {'recidx': recidx,
                                                                      'rec_fldtype': rec[fldType],
                                                                      'rec': rec,
                                                                      'OCC_TYPE_CONV': OCC_TYPE_CONV }))
        # check for unique booking
        if rec[BOOKING_FLD] is None:
            continue
        elif rec[BOOKING_FLD] in booking:
            errors.append('Field {} booking already exists: {}'.format(fldType,
                                                                       {'recidx': recidx,
                                                                        'orig_recidx': booking[rec[BOOKING_FLD]]['recidx'],
                                                                        'booking': rec[BOOKING_FLD],
                                                                        'rec': rec}))
        else:
            booking[rec[BOOKING_FLD]]=copy.deepcopy(rec)
            booking[rec[BOOKING_FLD]]['recidx']=recidx

            
    # return errors found
    return errors
            
def filtered_sorted_xlsaref(xlsaref, fldFirstNight, fldNights):
    """
    take in the list of dicts
    remove fields with fields not properly populated
    sort the output by fldFirstNight
    return this list of dicts

    :param xlsaref: (list of dicts) recodrds from that file from sheet "Listing"
    :param fldFirstNight: (str) column header of the first night date column
    :param fldNights: (str) column header for the field holding the int # of nights

    :return xlsaref: (list of dicts) - filtered/sorted list of dicts
    """

    # find records where first date and number of nights are filled in
    xlsaref = [x for x in xlsaref if x[fldFirstNight] and (x[fldNights] or x[fldNights] == 0)]

    # sort these records so they are in date order
    xlsaref = sorted(xlsaref, key=itemgetter(fldFirstNight))

    return xlsaref


def calc_revenue_per_day(xlsaref, fldNights):
    """
    Calculate the revenue per day which is revenue/days

    :uses REVTOTAL_FLD: (str) - field were revenue is stored
    :uses REVPERDAY_FLD: (str) - field where revenue per day is placed/saved to

    :param xlsaref: (list of dict) - the records from the xls
    :param fldNights: (str) - field that houses the # of nights this record represents

    :returns - updated xlsaref with the revenue per day calculated
    """
    for rec in xlsaref:
        if rec[fldNights] and rec[REVTOTAL_FLD]:
            rec[REVPERDAY_FLD] = int(float(rec[REVTOTAL_FLD])/float(rec[fldNights])*100)/100
        else:
            rec[REVPERDAY_FLD] = None

def find_max_value_per_booking_code(xlsaref):
    """
    step through the records and extract the code from alpha before "-" and then find the max value after that

    :param xlsaref: (list of dict) - the records from the xls
    
    :return max_values: (dict) - keyed by the code with value of the max value
    """
    max_values = dict()
    for rec in xlsaref:
        booking = rec.get(BOOKING_FLD)
        if not booking:
            # if not there or not populated skip
            continue
        bk_code, bk_value = booking.split('-')
        bk_code = bk_code.upper()
        if bk_code not in max_values:
            max_values[bk_code] = bk_value
        elif bk_value > max_values[bk_code]:
            max_values[bk_code] = bk_value

    return max_values


def assign_booking_code(rec, max_values):
    """
    using the type field - determine what code to use to populate the value
    use the max_values to find the next value - increment and assign
    and update max_values with the newly assigned value

    :param rec: (dict) a record of data
    :param max_values: (dict) max value of each code type 

    :updates rec: sets the BOOKING_FLD field with a value based on type and max_values
    """
    booking_code = OCC_TYPE_2_BOOKING_CODE[OCC_TYPE_CONV[rec['Type']][0]]
    if booking_code not in max_values:
        max_values[booking_code] = 0
    next_value = int(int(max_values[booking_code]) / 10 + 1) * 10
    rec[BOOKING_FLD] = '{}-{:05d}'.format(booking_code, next_value)
    max_values[booking_code] = next_value

    return booking_code


def insert_holds_on_reservation(rec, recidx, xlsaref, booking_code, max_values):
    """
    using the booking code - determine if the record prior was the hold associated with this record
    and if not - add the two hold records

    """
    if booking_code[:3] not in ('MLS', 'OWN'):
        return {}, {}

    # create the variable
    hldrecin = {k:'' for k in rec}
    hldrecout = {k:'' for k in rec}

    debug = False
    if rec[BOOKING_FLD] == 'MLS-00080' and False:
        logger.warning('MLS-00080')
        debug = True

    # DEBUGGING
    if debug:
        print('recidx:', recidx)
        print('len-xlsaref:', len(xlsaref))

    # clean before is the day before they shos
    clean_before_start = rec['First Night'] - ADD_ONE_DAY
    clean_before_end = rec['First Night']

    clean_after_start = rec['Checkout Day']
    clean_after_end = rec['Checkout Day'] + ADD_ONE_DAY

    # calculate the date prior to the reservation
    hold_start = rec['First Night'] - ADD_ONE_DAY
    hold_end = rec['Checkout Day'] + ADD_ONE_DAY

    
    # calc prior record - and if the record is the first record we don't have a prior record so set it empty
    prior_rec = xlsaref[recidx - 1] if recidx else {}
    post_rec = xlsaref[recidx + 1] if len(xlsaref) > recidx+1 else {}

    # create the default hold/cleaning reocrd that we will adjust
    hldrecin['Type'] = 'Hold - Clean'
    hldrecin['Nights'] = 1
    hldrecin['Source'] = 'MS-Add'
#    hldrecin['Managing'] = 'MS' # removed 2023-01-21
    hldrecin['First Night'] = clean_before_start
    hldrecin['Checkout Day'] = clean_before_end

    hldrecout['Type'] = 'Hold - Clean'
    hldrecout['Nights'] = 1
    hldrecout['Source'] = 'MS-Add'
#    hldrecout['Managing'] = 'MS' # removed 2023-01-21
    hldrecout['First Night'] = clean_after_start
    hldrecout['Checkout Day'] = clean_after_end

    if debug:
        logger.warning('Clean_before_start: %s', clean_before_start)
        logger.warning('Clean_before_end: %s', clean_before_end)
        logger.warning('Must start later than: %s', clean_before_end - ADD_ONE_WEEK)


    # PRE #
    prior_booking_code = prior_rec.get(BOOKING_FLD)
    if prior_booking_code:
        prior_booking_code = prior_booking_code[:3]

    # determine if we need a hold record - before this record
    if prior_booking_code in ("CLN", "HLD"):
        # 1 record earlier is a cleaning reocrd - see if it is close enough
        if prior_rec.get('First Night') <= clean_before_end and \
           prior_rec.get('First Night') > clean_before_end - ADD_TWO_WEEK:
            if prior_rec.get('Checkout Day') <= clean_before_end:
                logger.info('Prior record works as a cleaning day: %s',
                            {'recidx': recidx,
                             'prior_booking_code': prior_booking_code,
                             'rec': rec})

                # clear the value - prior record is it
                hldrecin = {}
            else:
                # end date on cleaning does not work but start date does
                logger.warning('Cleaning end record conflicts with appt: %s',
                               {'recidx': recidx,
                                'rec': rec,
                                'error': 'please fix cleaning record'})
                sys.exit(1)
        else:
            # start date does not work on this so insert a new start date

            # set values on this record
            # - just use what was already calculdated - day before cleaning
            logger.warning('using standard hold')
            pass
        
    else:
        # this is a reservation before
        if prior_rec.get('Checkout Day') == clean_before_end:
            if False:
                logger.info('PRE same day hldrecin: %s', pp.pformat(hldrecin))
                logger.info('Rec: %s', rec)
                sys.exit()
            
            # we need to do a same day cleaning
            hldrecin['Nights'] = 0
            hldrecin['First Night'] = clean_before_end
        else:
            logger.warning('using standard hold on complete')
            # else - we just use what was calculated


    if debug:
        logger.warning('recidx: %s', recidx)
        logger.warning('rec: \n%s', pp.pformat(rec))
        logger.warning('prior rec: \n%s', pp.pformat(prior_rec))
        logger.warning('hldrecin: \n%s', pp.pformat(hldrecin))


    # taken an action if we have a new record to create
    # 2024-02-07;kv - removed the insert/addition of this record
    if hldrecin or 0:
        logger.warning('setting booking_code on new cleaning record for record: %s', recidx)
        hold_booking_code = assign_booking_code(hldrecin, max_values)
    else:
        logger.warning('get next record - nothing updated on this record: %s', recidx)

    # ----------------------------------------
    
    # POST #
    post_booking_code = post_rec.get(BOOKING_FLD)
    if post_booking_code:
        post_booking_code = post_booking_code[:3]

    
    # determine if we need a hold record - before this record
    if post_booking_code in ("CLN", "HLD"):
        # record after this record is a cleaning record
        # see if it is close enough that we don't need another one
        if post_rec.get('First Night') >= clean_after_start and \
           post_rec.get('First Night') <= clean_after_start + ADD_ONE_WEEK:
            logger.info('Post record works as a cleaning day: %s',
                        {'recidx': recidx,
                         'rec': rec})

            # clear the value - prior record is it
            hldrecout = {}

        else:
            # start date does not work on this so insert a new start date

            # set values on this record
            # - just use what was already calculdated - day before cleaning
            logger.warning('using standard post hold')
            pass
        
    else:
        # this is a reservation after
        if post_rec.get('First Night') == clean_after_start:
            if False:
                logger.info('Post same day hldrecin: %s', pp.pformat(hldrecin))
                logger.info('Post same day hldrecout: %s', pp.pformat(hldrecout))
                logger.info('Rec: %s', rec)
                sys.exit()
            
            # we need to do a same day cleaning
            hldrecout['Nights'] = 0
            hldrecout['Checkout Day'] = clean_after_start
        else:
            logger.warning('using standard hold on complete')
            # else - we just use what was calculated


    if debug:
        logger.warning('recidx: %s', recidx)
        logger.warning('rec: \n%s', pp.pformat(rec))
        logger.warning('post rec: \n%s', pp.pformat(post_rec))
        logger.warning('hldrecout: \n%s', pp.pformat(hldrecout))


    # taken an action if we have a new record to create
    # 2024-02-07;kv - removed the insert/addition of this record
    if hldrecout or 0:
        logger.warning('setting booking_code on new cleaning record for record: %s', recidx)
        hold_booking_code = assign_booking_code(hldrecout, max_values)
    else:
        logger.warning('get next record - nothing updated on this record: %s', recidx)

    if debug:
        sys.exit()

    return hldrecin, hldrecout


def update_xlsaref_records(xlsaref):
    # find the max value for each code
    max_values = None

    # caputre new hold
    new_holds = list()
    
    # only calculate this if we need it
    max_values = find_max_value_per_booking_code(xlsaref)
    
    # first sweep through and populated records with code
    for recidx, rec in enumerate(xlsaref):
        if not rec[BOOKING_FLD]:
            booking_code = assign_booking_code(rec, max_values)
        else:
            booking_code = rec[BOOKING_FLD]


        # remove the holds in and out 2024-02-07;kv
        if False:
            # check to see if we return a new value
            newholdin, newholdout = insert_holds_on_reservation(rec, recidx, xlsaref, booking_code, max_values)
            if newholdin:
                new_holds.append(newholdin)
            if newholdout:
                new_holds.append(newholdout)

        # calculate the $/night
        if rec.get(REVTOTAL_FLD) and rec.get(STAYS_FLD):
            rec[REVPERDAY_FLD] = float(rec.get(REVTOTAL_FLD))/float(rec[STAYS_FLD])

    # DEBUGGING
    if False:
        logger.warning('new_holds: \n%s', pp.pformat(new_holds))
        # logger.warning('xlsaref: \n%s', pp.pformat(xlsaref))
        sys.exit()
        
    # add the new records into xlsaref and then sort them
    xlsaref.extend(new_holds)
    xlsaref = sorted(xlsaref, key=itemgetter('First Night', 'Checkout Day'))

    # DEBUGGING
    if debug:
        logger.warning('updated xlsaref: \n%s', pp.pformat(xlsaref))
        sys.exit()

    return xlsaref

def rewrite_file(xlsfile, xlsaref, fldFirstNight, fldNights, xlsdateflds):
    """
    Take a list of dicts defined in xlsaref
    Take the list, sort by start date and filter out blank records

    Populate a new xlsx spread sheet that is properly formatted
    - First sheet called "Listing" will be the full list
    - 12 sheets after that will the the records per month

    Save the input file to .BAK and then create a new version of the file

    :param xlsfile: (str) input filename
    :param xlsaref: (list of dicts) records read in from xlsfile from sheet "Listing"
    :param fldFirstNight: (str) column header of the first night date column
    :param fldNights: (str) column header for the field holding the int # of nights
    :param xlsdateflds: (list) of field names that are date fields

    :return bak_filename: (str) - filename we converted the input filename into
    """
    
    # Debugging
    if False:
        header_keys = ['' if 'blank' in x else x for x in list(xlsaref[0].keys())]
        dict_keys = [x for x in list(xlsaref[0].keys())]
        print('header_keys:', header_keys)
        print('dict_keys:', dict_keys)
        print('xlsaref[0]:', xlsaref[0])
        sys.exit()

    # move the file so we can output with the same filename
    fname, fext = os.path.splitext(xlsfile)
    new_fname = fname + '.bak'
    if os.path.exists(new_fname):
        os.remove(new_fname)
    os.rename(xlsfile, new_fname)

    # find records where first date and number of nights are filled in
    # sort these records so they are in date order
    xlsaref = filtered_sorted_xlsaref(xlsaref, fldFirstNight, fldNights)

    # calculate the revenue per sheet
    calc_revenue_per_day(xlsaref, fldNights)

    # create workbook for output
    wb = openpyxl.Workbook()

    # set the first sheet name
    ws = wb.active
    ws.title = SHEET_LISTING

    # pull out the keys
    header_keys = ['' if 'blank' in x else x for x in list(xlsaref[0].keys())]
    dict_keys = [x for x in list(xlsaref[0].keys())]

    # build out the listing sheet
    # header
    for colidx, key in enumerate(header_keys, start=1):
        a1 = ws.cell(row=1, column=colidx, value=key)
        a1.font = bold
        if key in COL_CENTERED:
            a1.alignment = fit_centered
        else:
            a1.alignment = fit
        if key in COL_WIDTH:
            ws.column_dimensions[get_column_letter(colidx)].width = COL_WIDTH[key]

    # data records
    for recidx, rec in enumerate(xlsaref, start=2):
        for colidx, key in enumerate(dict_keys, start=1):
            if key not in rec:
                continue
            a1 = ws.cell(row=recidx, column=colidx, value=rec[key])
            a1.font = regular
            if key in COL_CENTERED:
                a1.alignment = fit_centered
            else:
                a1.alignment = fit
            if key in xlsdateflds:
                a1.number_format = "MM/DD/YYYY"
            if key in COL_NUMBER_FLDS:
                a1.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

    # build out month oriented tabs
    for mon in range(1, 13):
        # create the new sheet
        ws = wb.create_sheet(title=MON_STRING[mon - 1])
        # header
        for colidx, key in enumerate(header_keys, start=1):
            a1 = ws.cell(row=1, column=colidx, value=key)
            a1.font = bold
            if key in COL_CENTERED:
                a1.alignment = fit_centered
            else:
                a1.alignment = fit
            if key in COL_WIDTH:
                ws.column_dimensions[get_column_letter(colidx)].width = COL_WIDTH[key]

        # extract out data for this sheet
        monaref = [x for x in xlsaref if x[fldFirstNight].month == mon]

        # data records
        for recidx, rec in enumerate(monaref, start=2):
            for colidx, key in enumerate(dict_keys, start=1):
                a1 = ws.cell(row=recidx, column=colidx, value=rec[key])
                a1.font = regular
                if key in COL_CENTERED:
                    a1.alignment = fit_centered
                else:
                    a1.alignment = fit
                if key in xlsdateflds:
                    a1.number_format = "MM/DD/YYYY"

    wb.save(xlsfile)

    return new_fname, xlsaref


def find_and_remove_dup_start_dates(xlsaref, fldFirstNight, fldNights):
    overlap = list()
    new_xlsaref = list()
    max_start_nights = {}
    for rec in xlsaref:
        if rec[fldFirstNight] in max_start_nights:
            max_start_nights[rec[fldFirstNight]] = max(int(rec[fldNights]), max_start_nights[rec[fldFirstNight]])
            overlap.append(rec[fldFirstNight])
        else:
            max_start_nights[rec[fldFirstNight]] = int(rec[fldNights])

    for rec in xlsaref:
        if int(rec[fldNights]) == max_start_nights[rec[fldFirstNight]]:
            new_xlsaref.append(rec)

    return new_xlsaref, overlap


# routine that reads the XLS, converts the data, and saves it to the output file
# Global variables used:
#    DATE_FMT - format string for the date string
#    ADD_ONE_DAY - delta date value that adds one day
#    OCC_TYPE_CONV - conversion of the occupytype 
#
def load_convert_save_file(xlsfile,
                           req_cols,
                           occupy_filename,
                           fldFirstNight,
                           fldNights,
                           fldType,
                           xlsdateflds,
                           fldLastNight,
                           pool_heater_allowed_filename,
                           debug=False):
    """
    Load convert and save the file - this is like a main_function()

    :param xlsfile: (str) - name of the source xlsx file
    :req_cols: (list) - list of column headers we must find in this file
    :param occupy_filename: (str) - name of the output file that houses the nights the villa is occupied
    :param fldFirstNight: (str) - column header that captures the first night the villa is occupied (date field)
    :param fldNights: (str) - column header that captures the number of nights the villa is occupied (int)
    :param xlsdateflds: (list of str) - column headers that are date fields that must be converted
    :param pool_allowed_heater_filename: (str) - name of the putput file that houses the days that pool ON is enabled
    :param debug: (bool) - when set, we run in debug mode.

    """
    # logging
    logger.info('Read in XLS:%s', xlsfile)

    # read in the XLS
    xlsaref = kvxls.readxls2list_findheader(xlsfile, req_cols=req_cols,
                                            optiondict={'dateflds': xlsdateflds, 'sheetname': 'Listing'}, debug=False)

    # debugging
    if debug:
        print('xlsaref[0] read in:', xlsaref[0])

    # validate the content and set the Booking field if needed
    # and insert records for cleaning if they are not in here.
    xlsaref = filtered_sorted_xlsaref(xlsaref, fldFirstNight, fldNights)

    # debugging
    if debug:
        print('xlsaref[0] filtered:', xlsaref[0])

    # validate records are right
    errors = validate_res_records(xlsaref, fldFirstNight, fldNights, fldLastNight, fldType)
    if errors:
        for x in errors:
            print(x)
        sys.exit(1)

    # now validate the file and fill in fields need filling and insert records needing inserting (holds)
    xlsaref = update_xlsaref_records(xlsaref)

    # DEBUGGING
    if False:
        pp.pprint(xlsaref)
        sys.exit(1)
    
    # reformat the file and get back the filename of the original file renamed
    # and return the list of records sorted by firstNight
    bak_fname, xlsaref = rewrite_file(xlsfile, xlsaref, fldFirstNight, fldNights, xlsdateflds)

    logger.info("Saved orig file, reformatted file: %s",
                {'orig_file': xlsfile,
                 'bak_file': bak_fname})

    # remove duplicates if any exist
    xlsaref, overlap = find_and_remove_dup_start_dates(xlsaref, fldFirstNight, fldNights)

    if overlap:
        logger.warning('Multiple records with same start night: %s',
                    {'overlap': overlap})

    # capture if the current renter is still there - their start date
    current_guest_start = None

    # get the current date
    now = datetime.datetime.now()
    yesterday = now - datetime.timedelta(days=1)

    # logging
    logger.info('Create occupancy file:%s', occupy_filename)

    # capture the dates that we need to output
    pool_heater_allowed_dates = []

    # create the output file and start the conversion/output process
    with open(occupy_filename, 'w') as t:
        # create the header to the file
        t.write('date,occtype\n')

        # set the first exit date to something that will NOT match
        exitdate = datetime.datetime(2019, 1, 1)

        # written out dates - so we don't write it twice
        already_written_date = list()

        # run through the records read in
        for rec in xlsaref:
            # process the record from the file - convert first night into a date variable
            # eventdate = datetime.datetime.strptime(rec[fldFirstNight], DATE_FMT)
            eventdate = rec[fldFirstNight]

            # add to the dictionary the datetime value just calculated
            rec['startdate'] = eventdate

            # for each night of stay - plus occ_type days to model the day the guest exits
            for cnt in range(int(rec[fldNights]) + OCC_TYPE_CONV[rec[fldType]][1]):
                # check to see if this date is pool heater enabled
                if POOL_FLD in rec and rec[POOL_FLD]:
                    if eventdate not in pool_heater_allowed_dates and eventdate >= yesterday:
                        pool_heater_allowed_dates.append(eventdate)

                # skip the date if this date matches the exit date of the prior
                if eventdate == exitdate:
                    # skip the date if this date matches the exit date of the prior
                    logger.info('Start date is same as the last guests exit date:skip record creation:%s', exitdate)
                else:
                    # convert the eventdate to a string
                    eventdate_str = datetime.datetime.strftime(eventdate, DATE_FMT)
                    # now make sure we have not already written this date
                    if eventdate_str not in already_written_date:
                        # output this value
                        t.write('%s,%s\n' % (eventdate_str, OCC_TYPE_CONV[rec[fldType]][0]))
                        # and add this date to the list of dates written
                        already_written_date.append(eventdate_str)
                    else:
                        logger.warning('Skipped record as date already written: %s',
                                       {'eventdate': eventdate,
                                        'rec': rec})

                    # set the exit date to the last date written out
                    exitdate = eventdate

                # capture the date prior to looping
                rec['exitdate'] = exitdate

                # add one day to this date and loop
                eventdate += ADD_ONE_DAY

            # capture current guest if it exists
            if rec['startdate'] < now and rec['exitdate'] > now:
                current_guest_start = rec['startdate']


    # create a file pointer to open this desired file
    if True or pool_heater_allowed_dates:
        # dump records out
        with open(pool_heater_allowed_filename, 'w') as phaf:
            # output each date in the MM/DD/YYYY format
            for pool_date in pool_heater_allowed_dates:
                phaf.write('%s\n' % (datetime.datetime.strftime(pool_date, DATE_FMT)))

        logger.info('Create pool allowed file:%s', pool_heater_allowed_filename)
    
    # return the BP file with modifications
    return xlsaref, current_guest_start


# routine that read in the history stays file and the current stays files and addes to
# the history file any dates from the current stays file that are in the past
#
def migrate_stays_to_history(occupy_filename, occupy_history_filename, fldDate, debug=False):
    # log that we are doing this work
    # load stay history
    if os.path.isfile(occupy_history_filename):
        logger.info('migrate_stays_to_history:load file:%s', occupy_history_filename)
        stays_history = villaecobee.load_villa_calendar(occupy_history_filename, fldDate, debug=False)
    else:
        logger.info('migrate_stays_to_history:file does not exist:%s', occupy_history_filename)
        stays_history = dict()

    # load current stay information
    stays = villaecobee.load_villa_calendar(occupy_filename, fldDate, debug=False)
    logger.info('migrate_stays_to_history:load file:%s', occupy_filename)

    # capture today
    today = datetime.datetime.today()

    # and capture the number of records add to the stay history
    records_added = 0

    # step through the stays file and look for past due dates
    for staydate in stays:
        if not staydate:
            # skip blanks
            continue
        if datetime.datetime.strptime(staydate, DATE_FMT) < today:
            # this date is in the past - see if this date is in the history already
            if staydate not in stays_history:
                # add this date to stays history
                stays_history[staydate] = stays[staydate]
                # increment the counter
                records_added += 1
                # debugging message
                logger.debug('migrate_stays_to_history:date added:%s', staydate)
            else:
                # debugging message
                logger.debug('migrate_stays_to_history:date skipped:%s', staydate)

    # loop through - now if we added records we need to save stay history data
    if records_added:
        logger.info('migrate_stays_to_history:records added to history:%d', records_added)
        kvcsv.writedict2csv(occupy_history_filename, stays_history)
    else:
        logger.info('migrate_stays_to_history:no records added to history')


# ---------------------------------------------------------------------------
if __name__ == '__main__':

    # capture the command line
    optiondict = kvutil.kv_parse_command_line(optiondictconfig, debug=False)

    # set variables based on what came form command line
    debug = optiondict['debug']

    # logging
    kvutil.loggingAppStart(logger, optiondict, kvutil.scriptinfo()['name'])

    # migrate stays to history
    migrate_stays_to_history(optiondict['occupy_filename'], optiondict['occupy_history_filename'],
                             optiondict['fldDate'], debug=False)

    # load and convert the XLS to create the TXT
    xlsaref, current_guest_start = load_convert_save_file(optiondict['xls_filename'],
                                                          COL_REQUIRED,
                                                          optiondict['occupy_filename'],
                                                          optiondict['fldFirstNight'],
                                                          optiondict['fldNights'],
                                                          optiondict['fldType'],
                                                          optiondict['xlsdateflds'],
                                                          optiondict['fldLastNight'],
                                                          optiondict['pool_heater_allowed_filename'],
                                                          debug=optiondict['debug'])

    # if set copy the occupy_filename to the alternate directory
    if optiondict['occupy_alt_dir']:
        logger.info("copying  %s to %s", optiondict['occupy_filename'], optiondict['occupy_alt_dir'])
        shutil.copy(optiondict['occupy_filename'], optiondict['occupy_alt_dir'])
        
    # if set copy the occupy_filename to the alternate directory
    if optiondict['pool_heater_allowed_alt_dir']:
        logger.info("copying  %s to %s", optiondict['pool_heater_allowed_filename'], optiondict['pool_heater_allowed_alt_dir'])
        shutil.copy(optiondict['pool_heater_allowed_filename'], optiondict['pool_heater_allowed_alt_dir'])
        
    # if the google calendar sync flag is set - sync
    if optiondict['calendarsync']:
        # determine the starting date for the run
        if optiondict['startback']:
            now = datetime.datetime.now() + datetime.timedelta(days=optiondict['startback'])
        elif optiondict['startdate']:
            now = optiondict['startdate']
        elif current_guest_start:
            now = current_guest_start
        else:
            now = datetime.datetime.now()
        # now update the calendar
        villacalendar.sync_villa_cal_with_bp_xls(xlsaref, now=now, debug=optiondict['debug'])

    # validate we can load this file after we created it
    logger.info('Load newly created file looking for problems: %s', optiondict['occupy_filename'])
    try:
        villaecobee.load_villa_calendar(optiondict['occupy_filename'], optiondict['fldDate'], debug=False)
    except:
        # logging
        logger.error('Error in newly created file:%s', optiondict['occupy_filename'])

        # display message
        print("ERROR:unable to load the newly created file:see error in line above")
        print('Please correct file:', optiondict['occupy_filename'])

    # validate that we can load the file after we created it
    logger.info('Load newly created file looking for problems: %s', optiondict['pool_heater_allowed_filename'])
    try:
        pool.read_pool_heater_allowable_file(optiondict['pool_heater_allowed_filename'])
    except:
        # logging
        logger.error('Error in newly created file:%s', optiondict['pool_heater_allowed_filename'])

        # display message
        print("ERROR:unable to load the newly created file:see error in line above")
        print('Please correct file:', optiondict['pool_heater_allowed_filename'])
    # validate we can load this file after we created it
    logger.info('Load newly created file looking for problems: %s', optiondict['occupy_filename'])
    try:
        villaecobee.load_villa_calendar(optiondict['occupy_filename'], optiondict['fldDate'], debug=False)
    except:
        # logging
        logger.error('Error in newly created file:%s', optiondict['occupy_filename'])

        # display message
        print("ERROR:unable to load the newly created file:see error in line above")
        print('Please correct file:', optiondict['occupy_filename'])


    # ALTERNATE LOCATION CHECKS
    if optiondict['occupy_alt_dir']:
        optiondict['occupy_filename'] = os.path.join(optiondict['occupy_alt_dir'], optiondict['occupy_filename'])
        
        # validate we can load this file after we created it
        logger.info('Load newly created file looking for problems: %s', optiondict['occupy_filename'])
        try:
            villaecobee.load_villa_calendar(optiondict['occupy_filename'], optiondict['fldDate'], debug=False)
        except:
            # logging
            logger.error('Error in newly created file:%s', optiondict['occupy_filename'])
            
            # display message
            print("ERROR:unable to load the newly created file:see error in line above")
            print('Please correct file:', optiondict['occupy_filename'])

    
    if optiondict['pool_heater_allowed_alt_dir']:
        optiondict['pool_heater_allowed_filename'] = os.path.join(optiondict['pool_heater_allowed_alt_dir'], optiondict['pool_heater_allowed_filename'])
        
        # validate that we can load the file after we created it
        logger.info('Load newly created file looking for problems: %s', optiondict['pool_heater_allowed_filename'])
        try:
            pool.read_pool_heater_allowable_file(optiondict['pool_heater_allowed_filename'])
        except:
            # logging
            logger.error('Error in newly created file:%s', optiondict['pool_heater_allowed_filename'])
            
            # display message
            print("ERROR:unable to load the newly created file:see error in line above")
            print('Please correct file:', optiondict['pool_heater_allowed_filename'])

# eof
